"""
MedInventory - Export Service
Generates Excel and PDF reports for apparecchi, manutenzioni, and scadenzario.
"""

import io
from datetime import datetime, timedelta


def export_apparecchi_excel(apparecchi, divisione_nome=''):
    """Export apparecchi list to Excel (xlsx) using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Apparecchi"

    # Styles
    header_font = Font(name='Inter', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    title = f"Inventario Apparecchi Elettromedicali"
    if divisione_nome:
        title += f" - {divisione_nome}"
    ws.merge_cells('A1:L1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Inter', bold=True, size=14, color='0EA5E9')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:L2')
    ws['A2'] = f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Inter', size=9, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = [
        'Matricola', 'Descrizione', 'N. Inventario', 'Marca', 'Modello',
        'Anno', 'Classificazione', 'Ubicazione', 'Stato', 'Fornitore',
        'IP', 'Divisione'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_idx, app in enumerate(apparecchi, 5):
        data = [
            app.get('matricola', ''),
            app.get('descrizione', ''),
            app.get('numero_inventario', ''),
            app.get('marca', ''),
            app.get('modello', ''),
            app.get('anno_fabbricazione', ''),
            app.get('classificazione', ''),
            app.get('ubicazione', ''),
            app.get('stato', ''),
            app.get('fornitore', ''),
            app.get('ip_address', ''),
            app.get('divisione_nome', ''),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value or '')
            cell.border = thin_border
            cell.font = Font(name='Inter', size=10)

    # Column widths
    from openpyxl.utils import get_column_letter
    widths = [18, 14, 14, 16, 20, 8, 14, 18, 14, 18, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Auto-filter
    ws.auto_filter.ref = f"A4:L{len(apparecchi) + 4}"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_manutenzioni_excel(manutenzioni, title_extra=''):
    """Export manutenzioni list to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Manutenzioni"

    header_font = Font(name='Inter', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    title = f"Registro Manutenzioni"
    if title_extra:
        title += f" - {title_extra}"
    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Inter', bold=True, size=14, color='0EA5E9')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Inter', size=9, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = [
        'Data', 'Apparecchio', 'Matricola', 'Tipo', 'Tecnico/Ditta',
        'Descrizione', 'Esito', 'Costo', 'Prossima Scadenza', 'Divisione'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, m in enumerate(manutenzioni, 5):
        data = [
            m.get('data_intervento', ''),
            f"{m.get('marca', '')} {m.get('modello', '')}",
            m.get('matricola', ''),
            m.get('tipo', ''),
            m.get('tecnico_ditta', ''),
            m.get('descrizione', ''),
            m.get('esito', ''),
            m.get('costo', ''),
            m.get('prossima_scadenza', ''),
            m.get('divisione_nome', ''),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value or '')
            cell.border = thin_border
            cell.font = Font(name='Inter', size=10)

    from openpyxl.utils import get_column_letter
    widths = [12, 22, 16, 14, 20, 30, 12, 10, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A4:J{len(manutenzioni) + 4}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_scadenzario_excel(scadenze, title_extra=''):
    """Export scadenzario to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Scadenzario"

    header_font = Font(name='Inter', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Priority colors
    priority_fills = {
        'scaduto': PatternFill(start_color='FEE2E2', fill_type='solid'),
        'urgente': PatternFill(start_color='FEE2E2', fill_type='solid'),
        'attenzione': PatternFill(start_color='FFEDD5', fill_type='solid'),
        'avviso': PatternFill(start_color='FEF9C3', fill_type='solid'),
        'ok': PatternFill(start_color='DCFCE7', fill_type='solid'),
    }

    title = "Scadenzario Manutenzioni"
    if title_extra:
        title += f" - {title_extra}"
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Inter', bold=True, size=14, color='0EA5E9')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Inter', size=9, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = [
        'Apparecchio', 'Matricola', 'Ubicazione', 'Tipo Manutenzione',
        'Prossima Scadenza', 'Giorni Rimasti', 'Priorita', 'Divisione'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row_idx, s in enumerate(scadenze, 5):
        priorita = s.get('priorita', 'ok')
        data = [
            f"{s.get('marca', '')} {s.get('modello', '')}",
            s.get('matricola', ''),
            s.get('ubicazione', ''),
            s.get('tipo_manutenzione', ''),
            s.get('prossima_scadenza', ''),
            s.get('giorni_rimasti', ''),
            priorita.upper(),
            s.get('divisione_nome', ''),
        ]
        fill = priority_fills.get(priorita, PatternFill())
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value or '')
            cell.border = thin_border
            cell.font = Font(name='Inter', size=10)
            cell.fill = fill

    from openpyxl.utils import get_column_letter
    widths = [24, 16, 18, 16, 16, 14, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_verifiche_excel(verifiche, title_extra=''):
    """Export verifiche di sicurezza elettrica to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Verifiche Elettriche"

    header_font = Font(name='Inter', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    esito_fills = {
        'positivo': PatternFill(start_color='DCFCE7', fill_type='solid'),
        'negativo': PatternFill(start_color='FEE2E2', fill_type='solid'),
        'con_riserva': PatternFill(start_color='FEF9C3', fill_type='solid'),
    }

    title = "Registro Verifiche di Sicurezza Elettrica"
    if title_extra:
        title += f" - {title_extra}"
    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Inter', bold=True, size=14, color='0EA5E9')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Inter', size=9, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = [
        'Data Verifica', 'Apparecchio', 'Matricola', 'Esito',
        'Tecnico/Ditta', 'Periodicità (gg)', 'Prossima Scadenza',
        'Note', 'Documento', 'Divisione'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, v in enumerate(verifiche, 5):
        esito = v.get('esito', '')
        data = [
            v.get('data_verifica', ''),
            f"{v.get('marca', '')} {v.get('modello', '')}",
            v.get('matricola', ''),
            esito,
            v.get('tecnico_ditta', ''),
            v.get('periodicita_giorni', ''),
            v.get('prossima_scadenza', ''),
            v.get('note', ''),
            'Sì' if v.get('documento_path') else 'No',
            v.get('divisione_nome', ''),
        ]
        fill = esito_fills.get(esito, PatternFill())
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value or '')
            cell.border = thin_border
            cell.font = Font(name='Inter', size=10)
            if col == 4:
                cell.fill = fill

    from openpyxl.utils import get_column_letter
    widths = [14, 24, 16, 14, 22, 14, 16, 30, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A4:J{len(verifiche) + 4}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_verifiche_pdf(verifiche, divisione_nome='', structure_name='', app_name='MedInventory'):
    """Export verifiche to PDF."""
    from fpdf import FPDF

    class PDF(FPDF):
        def header(self):
            self.set_font('Helvetica', 'B', 14)
            self.cell(0, 8, app_name, ln=True, align='C')
            self.set_font('Helvetica', '', 9)
            subtitle = 'Registro Verifiche di Sicurezza Elettrica'
            if divisione_nome:
                subtitle += f' - {divisione_nome}'
            if structure_name:
                subtitle = f'{structure_name} - {subtitle}'
            self.cell(0, 5, subtitle, ln=True, align='C')
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 5, f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align='C')
            self.ln(3)

        def footer(self):
            self.set_y(-15)
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 10, f'Pagina {self.page_no()}/{{nb}}', align='C')

    esito_colors = {
        'positivo': (220, 252, 231),
        'negativo': (254, 202, 202),
        'con_riserva': (254, 249, 195),
    }

    pdf = PDF('L', 'mm', 'A4')
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    pdf.set_fill_color(14, 165, 233)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)

    col_widths = [25, 45, 25, 22, 35, 18, 25, 55, 10]
    col_headers = ['Data', 'Apparecchio', 'Matricola', 'Esito',
                   'Tecnico/Ditta', 'Periodicità', 'Prossima Scad.', 'Note', 'Doc.']

    for w, h in zip(col_widths, col_headers):
        pdf.cell(w, 7, h, border=1, align='C', fill=True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 7)

    for v in verifiche:
        esito = v.get('esito', '')
        color = esito_colors.get(esito, (255, 255, 255))
        pdf.set_fill_color(*color)
        row_data = [
            str(v.get('data_verifica', '') or ''),
            f"{v.get('marca', '')} {v.get('modello', '')}",
            str(v.get('matricola', '') or ''),
            esito,
            str(v.get('tecnico_ditta', '') or ''),
            str(v.get('periodicita_giorni', '') or ''),
            str(v.get('prossima_scadenza', '') or ''),
            str(v.get('note', '') or ''),
            'Sì' if v.get('documento_path') else 'No',
        ]
        for w, val in zip(col_widths, row_data):
            display_val = val[:int(w / 2)] if len(val) > int(w / 2) else val
            pdf.cell(w, 6, display_val, border=1, fill=True)
        pdf.ln()

    pdf.ln(5)
    pdf.set_font('Helvetica', 'I', 9)
    positivi = sum(1 for v in verifiche if v.get('esito') == 'positivo')
    negativi = sum(1 for v in verifiche if v.get('esito') == 'negativo')
    pdf.cell(0, 5, f"Totale verifiche: {len(verifiche)} | Positive: {positivi} | Negative: {negativi}", ln=True)

    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return buffer


def _foglio_semplice(titolo, intestazioni, righe_valori):
    """Foglio piano con intestazione formattata e colonne dimensionate.

    Niente raggruppamenti: in Excel si filtra e si ordina da soli, quindi la
    divisione e' una colonna e non una sezione.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = titolo[:31]

    ws.cell(row=1, column=1, value=titolo).font = Font(bold=True, size=13)

    for colonna, etichetta in enumerate(intestazioni, start=1):
        cella = ws.cell(row=3, column=colonna, value=etichetta)
        cella.font = Font(bold=True, color='FFFFFF')
        cella.fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9',
                                 fill_type='solid')
        cella.alignment = Alignment(horizontal='center')

    for indice, valori in enumerate(righe_valori, start=4):
        for colonna, valore in enumerate(valori, start=1):
            ws.cell(row=indice, column=colonna, value=valore)

    for colonna, etichetta in enumerate(intestazioni, start=1):
        larghezza = max([len(str(etichetta))] +
                        [len(str(v[colonna - 1] or '')) for v in righe_valori] or [10])
        ws.column_dimensions[ws.cell(row=3, column=colonna).column_letter].width = \
            min(max(larghezza + 2, 12), 40)

    ws.auto_filter.ref = f"A3:{ws.cell(row=3, column=len(intestazioni)).column_letter}" \
                         f"{max(3, len(righe_valori) + 3)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def stampa_inventario_excel(righe, titolo):
    """Foglio con lo stesso inventario del PDF: stesse righe, stesso ordine."""
    intestazioni = ['Marca', 'Modello', 'Descrizione', 'Matricola',
                    'Ubicazione', 'Divisione']
    valori = [[r.get('marca'), r.get('modello'), r.get('descrizione'),
               r.get('matricola'), r.get('ubicazione'), r.get('divisione_nome')]
              for r in righe]
    return _foglio_semplice(titolo, intestazioni, valori)


def stampa_scadenze_excel(scadute, in_scadenza, titolo):
    """Foglio con le stesse scadenze del PDF: prima le scadute, poi le altre."""
    intestazioni = ['Stato', 'Marca', 'Modello', 'Matricola', 'Ubicazione',
                    'Divisione', 'Scadenza', 'Giorni']
    valori = []
    for stato, gruppo in (('Scaduta', scadute), ('In scadenza', in_scadenza)):
        for r in gruppo:
            valori.append([stato, r.get('marca'), r.get('modello'),
                           r.get('matricola'), r.get('ubicazione'),
                           r.get('divisione_nome'), r.get('prossima_scadenza'),
                           r.get('giorni_rimasti')])
    return _foglio_semplice(titolo, intestazioni, valori)


def genera_report_scadenze_pdf(struttura_id, output_path):
    """Report scadenze per l'invio automatico via email.

    Usa lo stesso motore delle stampe manuali: il PDF che arriva in posta e'
    identico a quello che si stampa dalla pagina Stampe, e resta un solo posto
    da mantenere quando la grafica cambia.
    """
    from models import query_all, query_one, percorso_logo_struttura
    from report_service import stampa_scadenze

    struttura = query_one("SELECT nome, logo_path FROM strutture WHERE id = ?",
                          (struttura_id,))
    # Un solo "oggi" per l'intero prospetto, come in stampe_bp.scadenze():
    # date('now') di SQLite e' in UTC, datetime.now() in ora locale, e nella
    # fascia notturna italiana i due non coincidono. Qui il rischio e' minore
    # (l'etichetta del periodo che indica un giorno diverso dal confine reale
    # delle query), ma la classe di difetto e' la stessa e costa una riga.
    oggi = datetime.now().date()
    fine = oggi + timedelta(days=30)

    # ps.tipo_manutenzione porta il genere di manutenzione (preventiva,
    # correttiva, ...) e per le verifiche il letterale 'verifica_elettrica':
    # e' quello che distingue le righe in un prospetto che, a differenza
    # delle stampe manuali, non e' filtrato per tipo_record.
    base = """SELECT ps.marca, ps.modello, ps.matricola, ps.ubicazione,
                     ps.prossima_scadenza, ps.giorni_rimasti,
                     ps.tipo_manutenzione, d.nome AS divisione_nome
              FROM prossime_scadenze ps
              JOIN apparecchi a ON a.id = ps.apparecchio_id
              LEFT JOIN divisioni d ON d.id = ps.divisione_id
              WHERE a.struttura_id = ?"""

    scadute = query_all(
        base + " AND ps.prossima_scadenza < ? ORDER BY ps.prossima_scadenza",
        (struttura_id, oggi.isoformat()))
    in_scadenza = query_all(
        base + " AND ps.prossima_scadenza >= ? AND ps.prossima_scadenza <= ?"
               " ORDER BY ps.prossima_scadenza",
        (struttura_id, oggi.isoformat(), fine.isoformat()))

    contesto = {
        'struttura_nome': (struttura or {}).get('nome') or 'MedInventory',
        'titolo': 'Scadenzario',
        'ambito': '',
        'logo_path': percorso_logo_struttura(struttura),
        'mostra_firma': False,
        'mostra_spunta': False,
        'mostra_tipo': True,
        'fine_periodo': fine.strftime('%d/%m/%Y'),
    }

    with open(output_path, 'wb') as f:
        f.write(stampa_scadenze(scadute, in_scadenza, contesto))


def export_apparecchi_pdf(apparecchi, divisione_nome='', structure_name='', app_name='MedInventory'):
    """Export apparecchi list to PDF using fpdf2."""
    from fpdf import FPDF

    class PDF(FPDF):
        def header(self):
            self.set_font('Helvetica', 'B', 14)
            self.cell(0, 8, app_name, ln=True, align='C')
            self.set_font('Helvetica', '', 9)
            subtitle = 'Inventario Apparecchi Elettromedicali'
            if divisione_nome:
                subtitle += f' - {divisione_nome}'
            if structure_name:
                subtitle = f'{structure_name} - {subtitle}'
            self.cell(0, 5, subtitle, ln=True, align='C')
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 5, f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align='C')
            self.ln(3)

        def footer(self):
            self.set_y(-15)
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 10, f'Pagina {self.page_no()}/{{nb}}', align='C')

    pdf = PDF('L', 'mm', 'A4')  # Landscape
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Table header
    pdf.set_fill_color(14, 165, 233)  # #0ea5e9
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)

    col_widths = [30, 20, 30, 30, 15, 15, 30, 20, 25, 30, 30]
    col_headers = ['Matricola', 'Descrizione', 'Marca', 'Modello', 'Anno', 'Class.',
                   'Ubicazione', 'Stato', 'Fornitore', 'IP', 'Divisione']

    for w, h in zip(col_widths, col_headers):
        pdf.cell(w, 7, h, border=1, align='C', fill=True)
    pdf.ln()

    # Table data
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 7)

    for app_item in apparecchi:
        row_data = [
            str(app_item.get('matricola', '') or ''),
            str(app_item.get('descrizione', '') or ''),
            str(app_item.get('marca', '') or ''),
            str(app_item.get('modello', '') or ''),
            str(app_item.get('anno_fabbricazione', '') or ''),
            str(app_item.get('classificazione', '') or ''),
            str(app_item.get('ubicazione', '') or ''),
            str(app_item.get('stato', '') or ''),
            str(app_item.get('fornitore', '') or ''),
            str(app_item.get('ip_address', '') or ''),
            str(app_item.get('divisione_nome', '') or ''),
        ]
        for w, val in zip(col_widths, row_data):
            # Truncate long values
            display_val = val[:int(w / 2)] if len(val) > int(w / 2) else val
            pdf.cell(w, 6, display_val, border=1)
        pdf.ln()

    # Summary
    pdf.ln(5)
    pdf.set_font('Helvetica', 'I', 9)
    pdf.cell(0, 5, f"Totale apparecchi: {len(apparecchi)}", ln=True)

    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return buffer


def export_scadenzario_pdf(scadenze, divisione_nome='', structure_name='', app_name='MedInventory'):
    """Export scadenzario to PDF."""
    from fpdf import FPDF

    class PDF(FPDF):
        def header(self):
            self.set_font('Helvetica', 'B', 14)
            self.cell(0, 8, app_name, ln=True, align='C')
            self.set_font('Helvetica', '', 9)
            subtitle = 'Scadenzario Manutenzioni'
            if divisione_nome:
                subtitle += f' - {divisione_nome}'
            if structure_name:
                subtitle = f'{structure_name} - {subtitle}'
            self.cell(0, 5, subtitle, ln=True, align='C')
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 5, f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align='C')
            self.ln(3)

        def footer(self):
            self.set_y(-15)
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 10, f'Pagina {self.page_no()}/{{nb}}', align='C')

    pdf = PDF('L', 'mm', 'A4')
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Priority colors (R, G, B)
    priority_colors = {
        'scaduto': (254, 202, 202),
        'urgente': (254, 202, 202),
        'attenzione': (255, 237, 213),
        'avviso': (254, 249, 195),
        'ok': (220, 252, 231),
    }

    # Header
    pdf.set_fill_color(14, 165, 233)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)

    col_widths = [50, 25, 30, 25, 25, 25, 20, 55]
    col_headers = ['Apparecchio', 'Matricola', 'Ubicazione', 'Tipo Man.',
                   'Scadenza', 'Giorni', 'Priorita', 'Divisione']

    for w, h in zip(col_widths, col_headers):
        pdf.cell(w, 7, h, border=1, align='C', fill=True)
    pdf.ln()

    # Data
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 7)

    for s in scadenze:
        priorita = s.get('priorita', 'ok')
        color = priority_colors.get(priorita, (255, 255, 255))
        pdf.set_fill_color(*color)

        row_data = [
            f"{s.get('marca', '')} {s.get('modello', '')}",
            str(s.get('matricola', '') or ''),
            str(s.get('ubicazione', '') or ''),
            str(s.get('tipo_manutenzione', '') or ''),
            str(s.get('prossima_scadenza', '') or ''),
            str(s.get('giorni_rimasti', '') or ''),
            priorita.upper(),
            str(s.get('divisione_nome', '') or ''),
        ]
        for w, val in zip(col_widths, row_data):
            display_val = val[:int(w / 2)] if len(val) > int(w / 2) else val
            pdf.cell(w, 6, display_val, border=1, fill=True)
        pdf.ln()

    # Summary
    pdf.ln(5)
    pdf.set_font('Helvetica', 'I', 9)
    scaduti = sum(1 for s in scadenze if s.get('priorita') == 'scaduto')
    urgenti = sum(1 for s in scadenze if s.get('priorita') == 'urgente')
    pdf.cell(0, 5, f"Totale scadenze: {len(scadenze)} | Scadute: {scaduti} | Urgenti: {urgenti}", ln=True)

    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return buffer


def genera_libretto_impianto(impianto_id, output_path):
    """Il libretto dell'impianto: anagrafica, componenti, documenti, piano,
    storico interventi.

    Servizio interno: non convalida lo scope. La tenant-isolation passa dalla
    rotta chiamante (impianti.libretto), che verifica l'accesso con
    models.impianto_accessibile() prima di invocare questa funzione — qui
    l'impianto_id arriva gia' verificato e le query sulle tabelle figlie
    restano chiavi su di esso.

    Dei documenti riporta i metadati, non i file: e' un indice di cosa esiste e
    chi l'ha emesso, non un archivio da stampare.
    """
    from models import query_one, query_all, percorso_logo_struttura
    from report_service import ReportPDF, testo_sicuro as _t

    impianto = query_one(
        """SELECT i.*, d.nome as divisione_nome, s.nome as struttura_nome,
                  s.logo_path as struttura_logo_path,
                  m.ragione_sociale as manutentore_nome
           FROM impianti i
           LEFT JOIN divisioni d ON d.id = i.divisione_id
           LEFT JOIN strutture s ON s.id = i.struttura_id
           LEFT JOIN manutentori m ON m.id = i.manutentore_id
           WHERE i.id = ?""", (impianto_id,))
    if not impianto:
        raise ValueError(f"Impianto {impianto_id} inesistente")

    componenti = query_all(
        "SELECT * FROM impianti_componenti WHERE impianto_id = ?"
        " ORDER BY descrizione", (impianto_id,))
    documenti = query_all(
        "SELECT * FROM impianti_documenti WHERE impianto_id = ?"
        " ORDER BY data_documento DESC, id DESC", (impianto_id,))
    piano = query_all(
        "SELECT * FROM impianti_scadenze WHERE impianto_id = ?"
        " ORDER BY attiva DESC, prossima_scadenza", (impianto_id,))
    interventi = query_all(
        """SELECT i.*, m.ragione_sociale as manutentore_nome
           FROM impianti_interventi i
           LEFT JOIN manutentori m ON m.id = i.manutentore_id
           WHERE i.impianto_id = ? ORDER BY i.data_intervento DESC""",
        (impianto_id,))

    contesto = {
        'struttura_nome': impianto['struttura_nome'] or 'MedInventory',
        'titolo': f"Libretto impianto - {impianto['nome']}",
        'ambito': impianto['divisione_nome'] or '',
        'logo_path': percorso_logo_struttura(
            {'logo_path': impianto['struttura_logo_path']}),
        'mostra_firma': False,
    }

    pdf = ReportPDF(contesto)
    pdf.add_page()

    def titolo(testo):
        pdf.ln(3)
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 7, _t(testo), new_x='LMARGIN', new_y='NEXT')
        pdf.set_font('Helvetica', '', 9)

    def riga(etichetta, valore):
        pdf.cell(45, 5, _t(etichetta), border=0)
        pdf.multi_cell(0, 5, _t(valore if valore not in (None, '') else '-'),
                        new_x='LMARGIN', new_y='NEXT')

    titolo('Anagrafica')
    riga('Tipo', impianto['tipo_custom'] or impianto['tipo'])
    riga('Stato', impianto['stato'])
    riga('Ubicazione', impianto['ubicazione'])
    riga('Identificativo', impianto['identificativo'])
    riga('Anno installazione', impianto['anno_installazione'])
    riga('Manutentore', impianto['manutentore_nome'])
    riga('Descrizione', impianto['descrizione'])

    titolo('Componenti')
    if componenti:
        for c in componenti:
            pdf.multi_cell(0, 5, _t(
                f"- {c['descrizione']} "
                f"({c['marca'] or '-'} {c['modello'] or ''} "
                f"mat. {c['matricola'] or '-'})"),
                new_x='LMARGIN', new_y='NEXT')
    else:
        pdf.cell(0, 5, _t('Nessun componente censito.'),
                 new_x='LMARGIN', new_y='NEXT')

    titolo('Documentazione')
    if documenti:
        for d in documenti:
            pdf.multi_cell(0, 5, _t(
                f"- [{d['tipo']}] {d['descrizione'] or d['filename']} - "
                f"{d['data_documento'] or 's.d.'} - "
                f"{d['emittente_ragione_sociale'] or 'emittente non indicato'}"),
                new_x='LMARGIN', new_y='NEXT')
    else:
        pdf.cell(0, 5, _t('Nessun documento caricato.'),
                 new_x='LMARGIN', new_y='NEXT')

    titolo('Piano di manutenzione e verifica')
    if piano:
        for p in piano:
            periodo = (f"ogni {p['periodicita_mesi']} mesi"
                       if p['periodicita_mesi'] else 'una tantum')
            stato = '' if p['attiva'] else ' [sospesa]'
            pdf.multi_cell(0, 5, _t(
                f"- {p['nome']} ({periodo}) - prossima: "
                f"{p['prossima_scadenza']} - "
                f"{p['riferimento_normativo'] or 'nessun riferimento'}{stato}"),
                new_x='LMARGIN', new_y='NEXT')
    else:
        pdf.cell(0, 5, _t('Piano non ancora definito.'),
                 new_x='LMARGIN', new_y='NEXT')

    titolo('Storico interventi')
    if interventi:
        for i in interventi:
            pdf.multi_cell(0, 5, _t(
                f"- {i['data_intervento']} [{i['tipo']}] "
                f"{i['esito'] or 'esito non indicato'} - "
                f"{i['manutentore_nome'] or i['tecnico_ditta'] or '-'} - "
                f"{i['descrizione'] or ''}"),
                new_x='LMARGIN', new_y='NEXT')
    else:
        pdf.cell(0, 5, _t('Nessun intervento registrato.'),
                 new_x='LMARGIN', new_y='NEXT')

    pdf.output(output_path)
    return output_path
