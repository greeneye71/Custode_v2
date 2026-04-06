"""
MedInventory - AI Service
Integration with AI models for:
1. Inventory import - parse Excel/PDF/CSV into structured device data
2. Email verbale parsing - extract maintenance data from PDF reports
3. Verifiche parsing - extract electrical safety verification data

Supported providers:
- Anthropic Claude (cloud API)
- Ollama (local, OpenAI-compatible)
- LM Studio (local, OpenAI-compatible)
- Any OpenAI-compatible endpoint
"""

import base64
import json
import csv
import io
import os
import logging

logger = logging.getLogger('medinventory.ai')

INVENTORY_SYSTEM_PROMPT = """Sei un assistente specializzato nell'analisi di inventari di apparecchi elettromedicali.
Ti verrà fornito il testo estratto da un documento (Excel, PDF o CSV) contenente un elenco di dispositivi medici.

Devi estrarre i dati e restituire un array JSON con i seguenti campi per ogni apparecchio trovato:
- matricola (obbligatorio, identificativo unico del dispositivo - vedi regole mappatura colonne)
- marca (obbligatorio)
- modello (obbligatorio)
- descrizione (opzionale)
- numero_inventario (opzionale)
- anno_fabbricazione (opzionale, solo anno numerico)
- classificazione (opzionale, uno tra: I, IIa, IIb, III)
- ubicazione (opzionale)
- fornitore (opzionale)
- codice_fornitore (opzionale, codice articolo o ricambio del fornitore)
- garanzia_scadenza (opzionale, data scadenza garanzia formato YYYY-MM-DD)
- contratto_manutenzione (opzionale, riferimento a contratto di manutenzione)
- ip_address (opzionale)
- note (opzionale)

REGOLE MAPPATURA COLONNE — segui questa priorità per identificare la matricola:
1. Colonne che indicano SEMPRE la matricola (numero di serie del costruttore):
   "Seriale", "Nr. Seriale", "N. Seriale", "Numero Seriale", "Numero di Serie",
   "Serial", "Serial Number", "Serial No", "Serial No.", "S/N", "SN",
   "Matricola", "Nr. Matricola", "N. Matricola",
   "Numero Serie", "No. Serie", "N° Serie", "N°Serie"
2. Se nessuna delle precedenti è presente, usa come fallback:
   "Numero di Inventario", "Nr. Inventario", "N. Inventario", "Inv.", "Inventario" → numero_inventario (non matricola)
3. Colonne "Codice", "Cod.", "Codice Interno", "Cod. Interno", "Cod. Int." → descrizione (NON matricola)
4. NON usare mai una colonna "Codice" come matricola se è presente una colonna seriale/matricola

ALTRE REGOLE:
- Restituisci SOLO un array JSON valido, senza altro testo
- Se un campo non è presente nel documento, omettilo o usa null
- Se non riesci a identificare la matricola in nessun modo, usa il valore più univoco disponibile
- Normalizza i nomi delle marche mantenendoli come nel documento (es. "GE" rimane "GE")
- Per la classificazione, mappa i valori comuni: "classe 1" -> "I", "IIA" -> "IIa", etc.
"""

VERBALE_SYSTEM_PROMPT = """Sei un assistente specializzato nell'analisi di verbali di manutenzione di apparecchi elettromedicali.
Ti verrà fornito il testo (o il PDF) di uno o più verbali/rapporti di intervento. Il documento può contenere interventi su più apparecchi.

Devi estrarre i dati e restituire un ARRAY JSON. Ogni elemento dell'array rappresenta un intervento su un singolo apparecchio e contiene:
- matricola (il numero di serie/matricola dell'apparecchio su cui è stato fatto l'intervento)
- tipo (uno tra: preventiva, correttiva, verifica, calibrazione)
- data_intervento (formato YYYY-MM-DD)
- tecnico_ditta (nome del tecnico e/o ditta che ha eseguito l'intervento)
- descrizione (descrizione sintetica dell'intervento)
- esito (esito: positivo, negativo, con riserva, etc.)
- costo (importo numerico, solo il numero senza simbolo euro, null se non presente)
- prossima_scadenza (data nel formato YYYY-MM-DD se indicata, null altrimenti)
- periodicita_giorni (periodicità in giorni se indicata, null altrimenti)

REGOLE:
- Restituisci SOLO un array JSON valido, senza altro testo
- Se il documento riguarda un solo apparecchio, restituisci un array con un solo elemento
- Se il documento riguarda più apparecchi (pagine diverse, sezioni diverse), restituisci un elemento per ciascuno
- La matricola è fondamentale per identificare l'apparecchio
- Per il tipo, deduci dal contesto: "manutenzione programmata" -> "preventiva", "guasto" -> "correttiva", etc.
- Se la data non è in formato standard, convertila in YYYY-MM-DD
- Se non trovi un campo, usa null
"""

CLASSIFICATION_SYSTEM_PROMPT = """Classifica il tipo di documento medico.
Rispondi SOLO con una di queste parole esatte:
- inventario — se il documento è un elenco/inventario di apparecchi elettromedicali (tabella con lista dispositivi, censimento, dotazione)
- verbale_manutenzione — se il documento è un verbale/rapporto di intervento di manutenzione (rapporto tecnico, intervento su guasto, manutenzione preventiva/correttiva)
- verifica_elettrica — se riguarda verifiche di sicurezza elettrica (IEC 62353, corrente di dispersione, messa a terra, VSE, collaudo elettrico)

Rispondi con UNA SOLA PAROLA tra: inventario, verbale_manutenzione, verifica_elettrica"""

VERIFICA_BATCH_SYSTEM_PROMPT = """Sei un assistente specializzato nell'analisi di rapporti di verifica di sicurezza elettrica per apparecchi elettromedicali.
Ti verrà fornito il testo estratto da un documento (PDF, Excel o CSV) che può contenere una o più verifiche di sicurezza elettrica.

Devi estrarre i dati e restituire un array JSON. Ogni elemento dell'array rappresenta una verifica:
- matricola (il numero di serie/matricola dell'apparecchio verificato - fondamentale)
- data_verifica (formato YYYY-MM-DD - obbligatorio)
- prossima_scadenza (formato YYYY-MM-DD - opzionale, calcolata dalla periodicità se non esplicitata)
- periodicita_giorni (365 per annuale, 730 per biennale - default 730 se non specificato)
- esito (uno tra: positivo, negativo, con_riserva - obbligatorio)
- tecnico_ditta (nome del tecnico e/o ditta che ha eseguito la verifica - opzionale)
- note (osservazioni o note del rapporto - opzionale)

REGOLE:
- Restituisci SOLO un array JSON valido, senza altro testo
- Se il documento contiene più apparecchi/verifiche, restituisci un elemento per ognuno
- La matricola è il campo più importante per identificare l'apparecchio
- Per l'esito: "pass", "OK", "conforme", "idoneo" → positivo; "fail", "KO", "non conforme", "non idoneo" → negativo; "con riserva", "condizionale" → con_riserva
- Se la data non è in formato standard, convertila in YYYY-MM-DD
- Keywords che identificano verifiche elettriche: sicurezza elettrica, corrente di dispersione, IEC 62353, messa a terra, CEI, VSE
"""


# ---------------------------------------------------------------------------
# Provider abstraction
# ---------------------------------------------------------------------------

_GEMINI_COMPLETIONS_URL = (
    'https://generativelanguage.googleapis.com/v1beta/openai/chat/completions'
)
_GEMINI_GENERATE_URL = (
    'https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent'
)
_OPENAI_API_URL = 'https://api.openai.com/v1/chat/completions'

# Model lists (used by struttura config UI and test-ai endpoint)
ANTHROPIC_MODELS = [
    ('claude-opus-4-6',            'Claude Opus 4.6 — Massima potenza'),
    ('claude-sonnet-4-6',          'Claude Sonnet 4.6 — Bilanciato (consigliato import)'),
    ('claude-sonnet-4-20250514',   'Claude Sonnet 4 — Maggio 2025'),
    ('claude-haiku-4-5-20251001',  'Claude Haiku 4.5 — Veloce (consigliato email)'),
    ('claude-3-5-sonnet-20241022', 'Claude 3.5 Sonnet — Ottobre 2024'),
    ('claude-3-haiku-20240307',    'Claude 3 Haiku — Legacy veloce'),
]

GEMINI_MODELS = [
    ('gemini-2.5-flash-preview-04-17', 'Gemini 2.5 Flash Preview — Più recente'),
    ('gemini-2.0-flash',               'Gemini 2.0 Flash — Veloce ($0.10/1M)'),
    ('gemini-1.5-flash',               'Gemini 1.5 Flash — Economico ($0.075/1M)'),
    ('gemini-1.5-flash-8b',            'Gemini 1.5 Flash-8B — Minimo ($0.037/1M)'),
    ('gemini-1.5-pro',                 'Gemini 1.5 Pro — Qualità superiore'),
]

OPENAI_MODELS = [
    ('gpt-4o-mini', 'GPT-4o mini — Bilanciato ($0.15/1M) — supporta PDF'),
    ('gpt-4o',      'GPT-4o — Massima qualità ($2.50/1M) — supporta PDF'),
]

AI_PROVIDERS = [
    ('anthropic',          'Anthropic Claude (Cloud)'),
    ('gemini',             'Google Gemini (Cloud) — da $0.037/1M token'),
    ('openai',             'OpenAI (Cloud) — da $0.15/1M token'),
    ('ollama',             'Ollama (Locale)'),
    ('lmstudio',           'LM Studio (Locale)'),
    ('openai_compatible',  'Altro endpoint OpenAI-compatibile'),
]

AI_PROVIDER_DEFAULTS = {
    'ollama':            'http://localhost:11434',
    'lmstudio':          'http://localhost:1234',
    'openai_compatible': 'http://localhost:8080',
}


def _get_ai_config(config=None, struttura_id=None):
    """Get AI provider configuration.
    When struttura_id is given: reads ONLY from struttura_config (no global config fallback for AI keys).
    When struttura_id is None: reads from global config.
    """
    if config is None:
        from flask import current_app
        config = current_app.config.get('APP_CONFIG', {})

    if struttura_id:
        from models import get_struttura_config as _gsc
        def _sc(key, default=''):
            val = _gsc(struttura_id, key)
            return val if val is not None else default
    else:
        def _sc(key, default=''):
            return config.get(key, default)

    return {
        'provider':       _sc('ai_provider', 'anthropic'),
        'api_key':        _sc('anthropic_api_key', ''),
        'gemini_api_key': _sc('gemini_api_key', ''),
        'openai_api_key': _sc('openai_api_key', ''),
        'model_import':   _sc('ai_import_model', 'claude-sonnet-4-20250514'),
        'model_email':    _sc('ai_email_model', 'claude-haiku-4-5-20251001'),
        'local_base_url': _sc('ai_local_base_url', 'http://localhost:11434'),
        'local_model':    _sc('ai_local_model', ''),
    }


def get_ai_config(struttura_id=None, config=None):
    """Public alias for _get_ai_config.
    When struttura_id is given: reads ONLY from struttura_config (no global fallback).
    When struttura_id is None: reads from global config.
    """
    return _get_ai_config(config=config, struttura_id=struttura_id)


def _call_anthropic(system_prompt, user_message, api_key, model, max_tokens=4096):
    """Call Anthropic Claude API."""
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model=model,
        max_tokens=max_tokens,
        system=system_prompt,
        messages=[{"role": "user", "content": user_message}]
    )
    return message.content[0].text.strip()


def _call_anthropic_with_pdf(system_prompt, user_text, pdf_path, api_key, model, max_tokens=4096):
    """Call Anthropic Claude API with a PDF document."""
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    pdf_data = _pdf_to_base64(pdf_path)
    message = client.messages.create(
        model=model,
        max_tokens=max_tokens,
        extra_headers={"anthropic-beta": "pdfs-2024-09-25"},
        system=system_prompt,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": pdf_data,
                    },
                },
                {"type": "text", "text": user_text},
            ],
        }],
    )
    return message.content[0].text.strip()


def _call_openai_compatible(system_prompt, user_message, base_url, model, max_tokens=4096):
    """Call an OpenAI-compatible API (Ollama, LM Studio, etc.)."""
    import httpx

    # Normalize base URL
    base_url = base_url.rstrip('/')
    if not base_url.endswith('/v1'):
        base_url += '/v1'

    url = f"{base_url}/chat/completions"

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_message},
        ],
        "max_tokens": max_tokens,
        "temperature": 0.1,
    }

    with httpx.Client(timeout=300.0) as client:
        response = client.post(url, json=payload)
        response.raise_for_status()
        data = response.json()

    return data["choices"][0]["message"]["content"].strip()


def _gemini_extract_text(data):
    """Estrae il testo dalla risposta nativa Gemini. Solleva ValueError se assente."""
    candidate = data.get("candidates", [{}])[0]
    content = candidate.get("content")
    if not content:
        finish = candidate.get("finishReason", "UNKNOWN")
        raise ValueError(f"Gemini non ha restituito contenuto (finishReason: {finish})")
    text = next((p["text"] for p in content.get("parts", []) if "text" in p), None)
    if not text:
        finish = candidate.get("finishReason", "UNKNOWN")
        raise ValueError(f"Gemini non ha restituito testo (finishReason: {finish})")
    return text.strip()


_GEMINI_TRANSIENT_ERRORS = {429, 500, 502, 503, 504}
_GEMINI_RETRY_DELAYS = [3, 10]  # secondi tra i tentativi


def _gemini_post(client, url, payload, api_key):
    """POST verso Gemini con retry automatico su errori transitori (503, 429, ecc.)."""
    import time
    last_exc = None
    for attempt, delay in enumerate([0] + _GEMINI_RETRY_DELAYS):
        if delay:
            time.sleep(delay)
        try:
            response = client.post(url, json=payload, params={"key": api_key})
            if response.status_code in _GEMINI_TRANSIENT_ERRORS and attempt < len(_GEMINI_RETRY_DELAYS):
                last_exc = None
                continue
            response.raise_for_status()
            return response.json()
        except Exception as e:
            last_exc = e
            if hasattr(e, 'response') and e.response.status_code not in _GEMINI_TRANSIENT_ERRORS:
                raise
    raise last_exc


def _call_gemini(system_prompt, user_message, api_key, model, max_tokens=4096):
    """Chiama l'API nativa Google Gemini (solo testo)."""
    import httpx

    url = _GEMINI_GENERATE_URL.format(model=model)
    payload = {
        "system_instruction": {"parts": [{"text": system_prompt}]},
        "contents": [{"parts": [{"text": user_message}]}],
        "generationConfig": {
            "maxOutputTokens": max_tokens,
            "temperature": 0.1,
        },
    }
    with httpx.Client(timeout=300.0) as client:
        data = _gemini_post(client, url, payload, api_key)

    return _gemini_extract_text(data)


def _call_gemini_with_pdf(system_prompt, user_text, pdf_path, api_key, model, max_tokens=4096):
    """Chiama l'API nativa Google Gemini con un PDF in-line base64."""
    import httpx

    pdf_data = _pdf_to_base64(pdf_path)
    url = _GEMINI_GENERATE_URL.format(model=model)
    payload = {
        "system_instruction": {"parts": [{"text": system_prompt}]},
        "contents": [{
            "parts": [
                {"inline_data": {"mime_type": "application/pdf", "data": pdf_data}},
                {"text": user_text},
            ]
        }],
        "generationConfig": {
            "maxOutputTokens": max_tokens,
            "temperature": 0.1,
        },
    }
    with httpx.Client(timeout=300.0) as client:
        data = _gemini_post(client, url, payload, api_key)

    return _gemini_extract_text(data)


def _call_openai(system_prompt, user_message, api_key, model, max_tokens=4096):
    """Chiama l'API OpenAI (cloud) — solo testo."""
    import httpx

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_message},
        ],
        "max_tokens": max_tokens,
        "temperature": 0.1,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type":  "application/json",
    }
    with httpx.Client(timeout=300.0) as client:
        response = client.post(_OPENAI_API_URL, json=payload, headers=headers)
        response.raise_for_status()
        data = response.json()

    return data["choices"][0]["message"]["content"].strip()


def _call_openai_with_pdf(system_prompt, user_text, pdf_path, api_key, model, max_tokens=4096):
    """Chiama l'API OpenAI con un PDF in-line base64 (supportato da GPT-4o e GPT-4o-mini)."""
    import httpx

    pdf_data = _pdf_to_base64(pdf_path)
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {
                        "type": "file",
                        "file": {
                            "filename": os.path.basename(pdf_path),
                            "file_data": f"data:application/pdf;base64,{pdf_data}",
                        },
                    },
                    {"type": "text", "text": user_text},
                ],
            },
        ],
        "max_tokens": max_tokens,
        "temperature": 0.1,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type":  "application/json",
    }
    with httpx.Client(timeout=300.0) as client:
        response = client.post(_OPENAI_API_URL, json=payload, headers=headers)
        response.raise_for_status()
        data = response.json()

    return data["choices"][0]["message"]["content"].strip()


def _call_ai(system_prompt, user_message, api_key, model, max_tokens=4096, config=None, struttura_id=None):
    """Unified AI call that routes to the correct provider."""
    ai_cfg = _get_ai_config(config, struttura_id)
    provider = ai_cfg['provider']

    if provider == 'anthropic':
        return _call_anthropic(system_prompt, user_message, api_key, model, max_tokens)
    elif provider == 'gemini':
        return _call_gemini(system_prompt, user_message, ai_cfg['gemini_api_key'], model, max_tokens)
    elif provider == 'openai':
        return _call_openai(system_prompt, user_message, ai_cfg['openai_api_key'], model, max_tokens)
    else:
        # ollama, lmstudio, openai_compatible
        base_url = ai_cfg['local_base_url']
        local_model = ai_cfg['local_model'] or model
        return _call_openai_compatible(system_prompt, user_message, base_url, local_model, max_tokens)


def _call_ai_with_pdf(system_prompt, user_text, pdf_path, api_key, model, max_tokens=4096, config=None, struttura_id=None):
    """Chiama l'AI con un documento PDF. Supporto nativo per Anthropic, Gemini e OpenAI."""
    ai_cfg = _get_ai_config(config, struttura_id)
    provider = ai_cfg['provider']

    if provider == 'anthropic':
        return _call_anthropic_with_pdf(system_prompt, user_text, pdf_path, api_key, model, max_tokens)
    elif provider == 'gemini':
        return _call_gemini_with_pdf(system_prompt, user_text, pdf_path, ai_cfg['gemini_api_key'], model, max_tokens)
    elif provider == 'openai':
        return _call_openai_with_pdf(system_prompt, user_text, pdf_path, ai_cfg['openai_api_key'], model, max_tokens)
    else:
        # Modelli locali: non supportano PDF nativamente — estrae il testo
        pdf_text = extract_from_pdf(pdf_path)
        if not pdf_text or len(pdf_text.strip()) < 20:
            raise ValueError(
                "Il PDF è scansionato (immagine) e il provider AI locale non supporta "
                "l'analisi diretta di PDF. Utilizzare Anthropic, Gemini o OpenAI."
            )
        combined = f"{user_text}\n\n{pdf_text[:15000]}"
        return _call_ai(system_prompt, combined, api_key, model, max_tokens, config, struttura_id)


def is_anthropic_provider(config=None, struttura_id=None):
    """Check if the current AI provider is Anthropic."""
    ai_cfg = _get_ai_config(config, struttura_id)
    return ai_cfg['provider'] == 'anthropic'


def check_ai_configured(config=None, struttura_id=None):
    """Check if AI is properly configured. Returns (ok, error_message)."""
    ai_cfg = _get_ai_config(config, struttura_id)
    provider = ai_cfg['provider']

    if provider == 'anthropic':
        if not ai_cfg['api_key']:
            return False, 'Chiave API Anthropic non configurata. Configura l\'AI nella pagina della struttura.'
        return True, None
    elif provider == 'gemini':
        if not ai_cfg['gemini_api_key']:
            return False, 'Chiave API Google Gemini non configurata. Configura l\'AI nella pagina della struttura.'
        return True, None
    elif provider == 'openai':
        if not ai_cfg['openai_api_key']:
            return False, 'Chiave API OpenAI non configurata. Configura l\'AI nella pagina della struttura.'
        return True, None
    else:
        if not ai_cfg['local_base_url']:
            return False, 'URL del server AI locale non configurato. Configura l\'AI nella pagina della struttura.'
        if not ai_cfg['local_model']:
            return False, 'Modello AI locale non configurato. Configura l\'AI nella pagina della struttura.'
        return True, None


# ---------------------------------------------------------------------------
# File text extraction
# ---------------------------------------------------------------------------

def extract_text_from_file(filepath, filetype):
    """Extract text content from Excel, PDF, or CSV files."""
    if filetype in ('xlsx', 'xls'):
        return _extract_from_excel(filepath)
    elif filetype == 'pdf':
        return extract_from_pdf(filepath)
    elif filetype == 'csv':
        return _extract_from_csv(filepath)
    else:
        raise ValueError(f"Tipo file non supportato: {filetype}")


def _extract_from_excel(filepath):
    """Extract text from Excel file using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"--- Foglio: {sheet.title} ---")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell) if cell is not None else '' for cell in row]
            if any(v.strip() for v in values):
                lines.append('\t'.join(values))
    wb.close()
    return '\n'.join(lines)


def extract_from_pdf(filepath):
    """Extract text from PDF using pdfplumber."""
    import pdfplumber
    text_parts = []
    try:
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                try:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
                except Exception:
                    pass
                try:
                    tables = page.extract_tables()
                    for table in tables:
                        for row in table:
                            if row:
                                values = [str(cell) if cell else '' for cell in row]
                                text_parts.append('\t'.join(values))
                except Exception:
                    pass
    except Exception:
        pass
    return '\n'.join(text_parts)


def _pdf_to_base64(filepath):
    with open(filepath, 'rb') as f:
        return base64.standard_b64encode(f.read()).decode('utf-8')


def _extract_from_csv(filepath):
    """Extract text from CSV file."""
    lines = []
    for encoding in ['utf-8', 'latin-1', 'cp1252']:
        try:
            with open(filepath, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    lines.append('\t'.join(row))
            break
        except UnicodeDecodeError:
            continue
    return '\n'.join(lines)


def get_pdf_page_count(filepath):
    """Get number of pages in a PDF."""
    import pdfplumber
    with pdfplumber.open(filepath) as pdf:
        return len(pdf.pages)


def extract_text_from_pdf_page(filepath, page_num):
    """Extract text from a specific PDF page (0-indexed)."""
    import pdfplumber
    with pdfplumber.open(filepath) as pdf:
        if page_num >= len(pdf.pages):
            return ''
        page = pdf.pages[page_num]
        parts = []
        text = page.extract_text()
        if text:
            parts.append(text)
        try:
            for table in (page.extract_tables() or []):
                for row in table:
                    if row:
                        parts.append('\t'.join(str(c) if c else '' for c in row))
        except Exception:
            pass
        return '\n'.join(parts)


def split_pdf_pages(filepath, output_dir):
    """Split a multi-page PDF into individual single-page PDF files.
    Returns list of output file paths.
    """
    from pypdf import PdfReader, PdfWriter

    os.makedirs(output_dir, exist_ok=True)
    reader = PdfReader(filepath)
    page_paths = []

    for i, page in enumerate(reader.pages):
        writer = PdfWriter()
        writer.add_page(page)
        page_path = os.path.join(output_dir, f"pagina_{i+1}.pdf")
        with open(page_path, 'wb') as f:
            writer.write(f)
        page_paths.append(page_path)

    return page_paths


# ---------------------------------------------------------------------------
# JSON parsing
# ---------------------------------------------------------------------------

def _parse_json_response(response_text, array=True):
    """Parse JSON from AI response, handling markdown fences and partial JSON."""
    try:
        return json.loads(response_text)
    except json.JSONDecodeError:
        clean = response_text
        if '```' in clean:
            import re
            m = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', clean)
            if m:
                clean = m.group(1)
        if array:
            start, end = clean.find('['), clean.rfind(']') + 1
        else:
            start, end = clean.find('{'), clean.rfind('}') + 1
        if start >= 0 and end > start:
            return json.loads(clean[start:end])
        raise ValueError(
            f"L'AI non ha restituito un JSON valido. "
            f"Risposta ({len(response_text)} caratteri): {response_text[:200]}..."
        )


def _parse_classification_result(result):
    """Parse document type from AI response using exact match before substring match."""
    r = result.lower().strip().rstrip('.')
    if r in ('verifica_elettrica', 'verifica'):
        return 'verifica_elettrica'
    if r in ('verbale_manutenzione', 'verbale'):
        return 'verbale_manutenzione'
    if r == 'inventario':
        return 'inventario'
    # Full-phrase substring fallback
    if 'verifica_elettrica' in r:
        return 'verifica_elettrica'
    if 'verbale_manutenzione' in r:
        return 'verbale_manutenzione'
    if 'inventario' in r:
        return 'inventario'
    return 'inventario'


# ---------------------------------------------------------------------------
# Inventory analysis
# ---------------------------------------------------------------------------

def analyze_inventory_with_ai(text, api_key, model='claude-sonnet-4-20250514', config=None, struttura_id=None):
    """Send extracted text to AI for structured parsing."""
    response_text = _call_ai(
        INVENTORY_SYSTEM_PROMPT,
        f"Analizza il seguente inventario ed estrai i dati degli apparecchi:\n\n{text[:15000]}",
        api_key, model, max_tokens=8192, config=config, struttura_id=struttura_id
    )
    items = _parse_json_response(response_text, array=True)
    if not isinstance(items, list):
        items = [items]
    return items, response_text


def analyze_inventory_from_pdf_document(filepath, api_key, model='claude-sonnet-4-20250514', config=None, struttura_id=None):
    """Analyze a scanned PDF inventory document."""
    response_text = _call_ai_with_pdf(
        INVENTORY_SYSTEM_PROMPT,
        "Analizza questo inventario ed estrai i dati degli apparecchi.",
        filepath, api_key, model, max_tokens=8192, config=config, struttura_id=struttura_id
    )
    items = _parse_json_response(response_text, array=True)
    if not isinstance(items, list):
        items = [items]
    return items, response_text


# ---------------------------------------------------------------------------
# Verbale (maintenance report) parsing
# ---------------------------------------------------------------------------

def parse_verbale_with_ai(pdf_text, api_key, model='claude-haiku-4-5-20251001', config=None, struttura_id=None):
    """Parse maintenance data from text."""
    response_text = _call_ai(
        VERBALE_SYSTEM_PROMPT,
        f"Analizza il seguente verbale di manutenzione:\n\n{pdf_text[:15000]}",
        api_key, model, max_tokens=4096, config=config, struttura_id=struttura_id
    )
    items = _parse_json_response(response_text, array=True)
    if not isinstance(items, list):
        items = [items]
    return items, response_text


def parse_verbale_from_pdf_document(filepath, api_key, model='claude-haiku-4-5-20251001', config=None, struttura_id=None):
    """Parse a scanned PDF maintenance report."""
    response_text = _call_ai_with_pdf(
        VERBALE_SYSTEM_PROMPT,
        "Analizza questo verbale di manutenzione ed estrai tutti gli interventi presenti.",
        filepath, api_key, model, max_tokens=4096, config=config, struttura_id=struttura_id
    )
    items = _parse_json_response(response_text, array=True)
    if not isinstance(items, list):
        items = [items]
    return items, response_text


# ---------------------------------------------------------------------------
# Verifiche (electrical safety) parsing
# ---------------------------------------------------------------------------

def analyze_verifiche_with_ai(text, api_key, model='claude-haiku-4-5-20251001', config=None, struttura_id=None):
    """Extract electrical safety verifications from text."""
    response_text = _call_ai(
        VERIFICA_BATCH_SYSTEM_PROMPT,
        f"Analizza il seguente documento ed estrai i dati delle verifiche di sicurezza elettrica:\n\n{text[:15000]}",
        api_key, model, max_tokens=4096, config=config, struttura_id=struttura_id
    )
    items = _parse_json_response(response_text, array=True)
    if not isinstance(items, list):
        items = [items]
    return items, response_text


def analyze_verifiche_from_pdf_document(filepath, api_key, model='claude-haiku-4-5-20251001', config=None, struttura_id=None):
    """Analyze a scanned PDF electrical safety report."""
    response_text = _call_ai_with_pdf(
        VERIFICA_BATCH_SYSTEM_PROMPT,
        "Analizza questo rapporto ed estrai i dati delle verifiche di sicurezza elettrica.",
        filepath, api_key, model, max_tokens=4096, config=config, struttura_id=struttura_id
    )
    items = _parse_json_response(response_text, array=True)
    if not isinstance(items, list):
        items = [items]
    return items, response_text


# ---------------------------------------------------------------------------
# Email document classification
# ---------------------------------------------------------------------------

def classify_email_document_type(pdf_text, api_key, model='claude-haiku-4-5-20251001', config=None, struttura_id=None):
    """Classify document type from PDF text.
    Returns: 'verifica_elettrica' | 'manutenzione'
    """
    text_lower = pdf_text.lower()

    # Keyword scoring per verifiche di sicurezza elettrica
    verifica_keywords = [
        'sicurezza elettrica', 'corrente di dispersione', 'iec 62353',
        'messa a terra', 'norma cei', 'vse', 'leakage current',
        'earth continuity', 'verifica elettrica', 'collaudo elettrico',
    ]
    score = sum(1 for kw in verifica_keywords if kw in text_lower)
    if score >= 2:
        return 'verifica_elettrica'

    manut_keywords = [
        'manutenzione', 'intervento', 'riparazione', 'sostituzione',
        'preventiva', 'correttiva', 'calibrazione', 'guasto',
    ]
    manut_score = sum(1 for kw in manut_keywords if kw in text_lower)
    if manut_score >= 2 and score == 0:
        return 'manutenzione'

    # Ambiguous: use AI
    try:
        result = _call_ai(
            "Rispondi solo con una parola: 'verifica_elettrica' se il documento riguarda verifiche di sicurezza elettrica (IEC 62353, corrente di dispersione, messa a terra), oppure 'manutenzione' negli altri casi.",
            f"Classifica questo documento:\n\n{pdf_text[:2000]}",
            api_key, model, max_tokens=20, config=config, struttura_id=struttura_id
        )
        if 'verifica' in result.lower():
            return 'verifica_elettrica'
    except Exception:
        pass
    return 'manutenzione'


def classify_email_document_type_from_pdf_document(filepath, api_key, model='claude-haiku-4-5-20251001', config=None, struttura_id=None):
    """Classify document type for a scanned PDF.
    Returns: 'verifica_elettrica' | 'manutenzione'
    """
    try:
        result = _call_ai_with_pdf(
            "Rispondi solo con una parola: 'verifica_elettrica' se il documento riguarda verifiche di sicurezza elettrica (IEC 62353, corrente di dispersione, messa a terra), oppure 'manutenzione' negli altri casi.",
            "Classifica questo documento.",
            filepath, api_key, model, max_tokens=200, config=config, struttura_id=struttura_id
        )
        if 'verifica' in result.lower():
            return 'verifica_elettrica'
    except Exception:
        pass
    return 'manutenzione'


# ---------------------------------------------------------------------------
# Unified document classification (import)
# ---------------------------------------------------------------------------

def classify_document_type(text, api_key, model='claude-haiku-4-5-20251001', config=None, struttura_id=None):
    """Classify document type from extracted text.
    Returns: 'inventario' | 'verbale_manutenzione' | 'verifica_elettrica'
    """
    text_lower = text[:5000].lower()

    verifica_kw = [
        'sicurezza elettrica', 'corrente di dispersione', 'iec 62353',
        'messa a terra', 'norma cei', 'vse', 'verifica elettrica',
        'collaudo elettrico', 'leakage current',
    ]
    v_score = sum(1 for kw in verifica_kw if kw in text_lower)

    manut_kw = [
        'verbale', 'rapporto di intervento', 'manutenzione preventiva',
        'manutenzione correttiva', 'rapporto tecnico', 'intervento tecnico',
        'riparazione', 'guasto',
    ]
    m_score = sum(1 for kw in manut_kw if kw in text_lower)

    inv_kw = [
        'inventario', 'elenco apparecchi', 'lista dispositivi',
        'censimento', 'dotazione',
    ]
    i_score = sum(1 for kw in inv_kw if kw in text_lower)

    if v_score >= 2:
        return 'verifica_elettrica'
    if m_score >= 2 and v_score == 0:
        return 'verbale_manutenzione'
    if i_score >= 1 and m_score == 0 and v_score == 0:
        return 'inventario'

    # Ambiguous: use AI — exception propagates to caller
    result = _call_ai(
        CLASSIFICATION_SYSTEM_PROMPT,
        f"Classifica questo documento:\n\n{text[:3000]}",
        api_key, model, max_tokens=200, config=config, struttura_id=struttura_id
    )
    return _parse_classification_result(result)


def classify_document_type_from_pdf(filepath, api_key, model='claude-haiku-4-5-20251001', config=None, struttura_id=None):
    """Classify document type from a scanned PDF.
    Returns: 'inventario' | 'verbale_manutenzione' | 'verifica_elettrica'
    """
    result = _call_ai_with_pdf(
        CLASSIFICATION_SYSTEM_PROMPT,
        "Classifica questo documento.",
        filepath, api_key, model, max_tokens=200, config=config, struttura_id=struttura_id
    )
    return _parse_classification_result(result)


# ---------------------------------------------------------------------------
# Duplicate detection
# ---------------------------------------------------------------------------

def find_duplicates(items, divisione_id, struttura_id=None):
    """Match extracted items against existing apparecchi in the database.
    Filtra per struttura_id quando disponibile per garantire isolamento multi-tenant.
    """
    from models import query_one, query_all

    results = []
    for item in items:
        result = {
            'data': item,
            'match_type': 'nuovo',
            'match_id': None,
            'match_confidence': 0,
            'match_info': None,
        }

        matricola = item.get('matricola', '').strip()
        descrizione = item.get('descrizione', '').strip() if item.get('descrizione') else ''

        if matricola:
            if struttura_id:
                existing = query_one(
                    "SELECT * FROM apparecchi WHERE matricola = ? AND struttura_id = ? AND stato != 'dismesso'",
                    (matricola, struttura_id)
                )
            else:
                existing = query_one(
                    "SELECT * FROM apparecchi WHERE matricola = ? AND stato != 'dismesso'",
                    (matricola,)
                )
            if existing:
                result['match_type'] = 'esatto'
                result['match_id'] = existing['id']
                result['match_confidence'] = 1.0
                result['match_info'] = f"{existing['marca']} {existing['modello']}"
                results.append(result)
                continue

        if descrizione:
            if struttura_id:
                existing = query_one(
                    "SELECT * FROM apparecchi WHERE descrizione = ? AND struttura_id = ? AND stato != 'dismesso'",
                    (descrizione, struttura_id)
                )
            else:
                existing = query_one(
                    "SELECT * FROM apparecchi WHERE descrizione = ? AND stato != 'dismesso'",
                    (descrizione,)
                )
            if existing:
                result['match_type'] = 'fuzzy'
                result['match_id'] = existing['id']
                result['match_confidence'] = 0.5
                result['match_info'] = f"{existing['marca']} {existing['modello']}"
                results.append(result)
                continue

        marca = item.get('marca', '').strip()
        modello = item.get('modello', '').strip()
        if marca and modello:
            if struttura_id:
                existing = query_one(
                    """SELECT * FROM apparecchi
                       WHERE LOWER(marca) = LOWER(?) AND LOWER(modello) = LOWER(?)
                       AND divisione_id = ? AND struttura_id = ?""",
                    (marca, modello, divisione_id, struttura_id)
                )
            else:
                existing = query_one(
                    """SELECT * FROM apparecchi
                       WHERE LOWER(marca) = LOWER(?) AND LOWER(modello) = LOWER(?)
                       AND divisione_id = ?""",
                    (marca, modello, divisione_id)
                )
            if existing:
                result['match_type'] = 'fuzzy'
                result['match_id'] = existing['id']
                result['match_confidence'] = 0.7
                result['match_info'] = f"{existing['marca']} {existing['modello']} ({existing['matricola']})"

        results.append(result)

    return results
