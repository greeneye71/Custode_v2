"""Test delle rotte di stampa: contano soprattutto i confini di visibilita'."""
import io
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader
from werkzeug.security import generate_password_hash


def testo_di(pdf_bytes):
    """Estrae il testo di tutte le pagine di un PDF (vedi test_report_service.py)."""
    lettore = PdfReader(io.BytesIO(pdf_bytes))
    return "\n".join(pagina.extract_text() for pagina in lettore.pages)


def _assert_rifiuto_pulito(client, url, messaggio):
    """Verifica che l'accesso fuori scope sia un redirect esplicito verso
    /stampe con un flash che spiega il motivo, non un 404/500 mascherato da
    un body che 'per caso' non comincia con %PDF."""
    risposta = client.get(url)
    assert risposta.status_code == 302
    assert risposta.headers['Location'] in ('/stampe', '/stampe/')

    pagina = client.get(url, follow_redirects=True)
    assert messaggio in pagina.get_data(as_text=True)


@pytest.fixture
def dati(app):
    """Due strutture, tre divisioni nella prima (una disattivata), un utente
    assegnato a una sola. Serve a coprire due confini distinti del filtro
    'tutte le divisioni' per admin/tecnico/superadmin (a.struttura_id = ?):
    che non sconfini nella struttura B (dx/ALT-1) e che includa anche le
    divisioni disattivate della propria struttura (d3/DIS-DIV-1), che
    restano fuori da g.divisioni ma sono comunque nell'ambito della
    struttura."""
    from models import execute
    with app.app_context():
        s1 = execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Clinica A','A',1)").lastrowid
        s2 = execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Clinica B','B',1)").lastrowid
        d1 = execute("INSERT INTO divisioni (nome,codice,struttura_id) VALUES ('Oculistica','OCU',?)", (s1,)).lastrowid
        d2 = execute("INSERT INTO divisioni (nome,codice,struttura_id) VALUES ('Cardiologia','CAR',?)", (s1,)).lastrowid
        d3 = execute("INSERT INTO divisioni (nome,codice,struttura_id,attiva) VALUES ('Dismessa','DIS',?,0)", (s1,)).lastrowid
        dx = execute("INSERT INTO divisioni (nome,codice,struttura_id) VALUES ('Altrui','ALT',?)", (s2,)).lastrowid
        hash_pw = generate_password_hash('Passw0rd!')
        execute(
            "INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,struttura_id,primo_accesso) "
            "VALUES ('admin@a.it',?,'A','A','admin',?,0)", (hash_pw, s1))
        utente = execute(
            "INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,struttura_id,primo_accesso) "
            "VALUES ('utente@a.it',?,'U','U','utente',?,0)", (hash_pw, s1)).lastrowid
        execute("INSERT INTO utenti_divisioni (utente_id,divisione_id,ruolo_divisione) VALUES (?,?,'utente')",
                (utente, d1))
        execute("INSERT INTO apparecchi (divisione_id,struttura_id,matricola,marca,modello,stato,ubicazione,descrizione) "
                "VALUES (?,?,'OCU-1','REXXAM','OZY','funzionante','Sala 1','Autorefrattometro')", (d1, s1))
        execute("INSERT INTO apparecchi (divisione_id,struttura_id,matricola,marca,modello,stato,ubicazione) "
                "VALUES (?,?,'CAR-1','GE','B40','funzionante','Sala 2')", (d2, s1))
        execute("INSERT INTO apparecchi (divisione_id,struttura_id,matricola,marca,modello,stato,ubicazione) "
                "VALUES (?,?,'DIS-DIV-1','DRAGER','X1','funzionante','Magazzino')", (d3, s1))
        execute("INSERT INTO apparecchi (divisione_id,struttura_id,matricola,marca,modello,stato,ubicazione) "
                "VALUES (?,?,'ALT-1','SIEMENS','Y1','funzionante','Sala X')", (dx, s2))
    return {'s1': s1, 's2': s2, 'd1': d1, 'd2': d2, 'd3': d3, 'dx': dx}


def entra(client, email):
    client.post('/login', data={'email': email, 'password': 'Passw0rd!'})


def test_la_pagina_stampe_risponde(client, dati):
    entra(client, 'admin@a.it')
    risposta = client.get('/stampe')
    assert risposta.status_code == 200


def test_la_pagina_stampe_risponde_anche_con_slash_finale(client, dati):
    entra(client, 'admin@a.it')
    risposta = client.get('/stampe/')
    assert risposta.status_code == 200


def test_admin_ottiene_l_inventario_di_struttura(client, dati):
    entra(client, 'admin@a.it')
    risposta = client.get('/stampe/inventario?divisione_id=tutte')
    assert risposta.status_code == 200
    assert risposta.data.startswith(b'%PDF')


def test_la_descrizione_arriva_dal_database_al_prospetto(client, dati):
    """La colonna esiste nel motore, ma serve a qualcosa solo se la query la
    chiede: senza a.descrizione nella SELECT il PDF esce con la colonna
    intestata e vuota, e il difetto non si vede da un controllo su %PDF."""
    entra(client, 'admin@a.it')
    testo = testo_di(client.get('/stampe/inventario?divisione_id=tutte').data)
    assert 'Descrizione' in testo          # intestazione di colonna
    assert 'Autorefrattometro' in testo    # valore dell'apparecchio OCU-1


def test_utente_non_ottiene_una_divisione_non_sua(client, dati):
    entra(client, 'utente@a.it')
    _assert_rifiuto_pulito(client, f"/stampe/inventario?divisione_id={dati['d2']}",
                            'Divisione non disponibile.')


def test_nessuno_ottiene_una_divisione_di_un_altra_struttura(client, dati):
    entra(client, 'admin@a.it')
    _assert_rifiuto_pulito(client, f"/stampe/inventario?divisione_id={dati['dx']}",
                            'Divisione non disponibile.')


def test_divisione_inesistente_non_produce_un_pdf(client, dati):
    entra(client, 'admin@a.it')
    _assert_rifiuto_pulito(client, '/stampe/inventario?divisione_id=999999',
                            'Divisione non disponibile.')


def test_utente_vede_solo_la_propria_divisione_nell_inventario_generale(client, dati):
    """Garanzia centrale: per il ruolo 'utente', 'tutte le divisioni' significa
    solo le sue, non l'intera struttura ne', a maggior ragione, una divisione
    disattivata a cui non e' assegnato."""
    entra(client, 'utente@a.it')
    risposta = client.get('/stampe/inventario?divisione_id=tutte')
    assert risposta.status_code == 200
    testo = testo_di(risposta.data)
    assert 'OCU-1' in testo
    assert 'CAR-1' not in testo
    assert 'DIS-DIV-1' not in testo


def test_admin_vede_tutte_le_divisioni_nell_inventario_generale(client, dati):
    """Simmetrico al precedente: per admin/tecnico/superadmin 'tutte le
    divisioni' significa l'intera struttura (a.struttura_id = ?), non
    l'elenco di g.divisioni. Percio' deve includere anche una divisione
    disattivata della propria struttura (DIS-DIV-1, che non compare in
    g.divisioni ma resta nell'ambito 'struttura'), e deve fermarsi al
    confine della struttura: un apparecchio di un'altra struttura (ALT-1)
    non deve mai comparire."""
    entra(client, 'admin@a.it')
    risposta = client.get('/stampe/inventario?divisione_id=tutte')
    assert risposta.status_code == 200
    testo = testo_di(risposta.data)
    assert 'OCU-1' in testo
    assert 'CAR-1' in testo
    assert 'DIS-DIV-1' in testo
    assert 'ALT-1' not in testo


def test_superadmin_senza_struttura_riceve_una_spiegazione(client, app, dati):
    """Senza struttura impersonata non c'e' un ambito su cui stampare: la pagina
    deve dirlo, non generare un PDF vuoto."""
    from models import execute
    from werkzeug.security import generate_password_hash
    with app.app_context():
        execute("INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,"
                "struttura_id,primo_accesso) VALUES ('super@a.it',?,'S','S',"
                "'superadmin',NULL,0)", (generate_password_hash('Passw0rd!'),))
    entra(client, 'super@a.it')

    pagina = client.get('/stampe', follow_redirects=True)
    assert 'contesto di una struttura' in pagina.get_data(as_text=True)

    _assert_rifiuto_pulito(client, '/stampe/inventario?divisione_id=tutte',
                            'Nessuna divisione accessibile.')


def test_scadenze_manutenzioni_produce_un_pdf(client, dati):
    entra(client, 'admin@a.it')
    risposta = client.get('/stampe/scadenze/manutenzioni?divisione_id=tutte&periodo=30g')
    assert risposta.status_code == 200
    assert risposta.data.startswith(b'%PDF')


def test_scadenze_verifiche_produce_un_pdf(client, dati):
    entra(client, 'admin@a.it')
    risposta = client.get('/stampe/scadenze/verifiche?divisione_id=tutte&periodo=anno')
    assert risposta.data.startswith(b'%PDF')


def test_tipo_di_scadenza_non_riconosciuto_viene_respinto(client, dati):
    entra(client, 'admin@a.it')
    _assert_rifiuto_pulito(client, '/stampe/scadenze/inventato?divisione_id=tutte&periodo=30g',
                            'Tipo di prospetto non riconosciuto.')


def test_date_incoerenti_non_producono_un_pdf(client, dati):
    entra(client, 'admin@a.it')
    _assert_rifiuto_pulito(
        client,
        '/stampe/scadenze/manutenzioni?divisione_id=tutte&periodo=date&da=2026-12-31&a=2026-01-01',
        'La data finale precede quella iniziale.')


def test_data_malformata_non_produce_un_pdf(client, dati):
    entra(client, 'admin@a.it')
    _assert_rifiuto_pulito(
        client,
        '/stampe/scadenze/manutenzioni?divisione_id=tutte&periodo=date&da=non-una-data&a=2026-01-01',
        'Indica due date valide nel formato anno-mese-giorno.')


def test_utente_non_stampa_le_scadenze_di_una_divisione_non_sua(client, dati):
    entra(client, 'utente@a.it')
    _assert_rifiuto_pulito(
        client, f"/stampe/scadenze/manutenzioni?divisione_id={dati['d2']}&periodo=30g",
        'Divisione non disponibile.')


def test_periodo_determina_quali_scadenze_rientrano_nel_prospetto(client, app, dati):
    """Una scadenza fra 45 giorni non deve comparire nel prospetto a 30 giorni,
    ma deve comparire in quello a 90 giorni. Un'asserzione che si limitasse a
    controllare che la risposta e' un PDF non si accorgerebbe di un bug in cui
    il periodo scelto viene ignorato dalla query."""
    from models import execute
    with app.app_context():
        execute(
            "INSERT INTO manutenzioni (apparecchio_id,tipo,data_intervento,prossima_scadenza) "
            "VALUES ((SELECT id FROM apparecchi WHERE matricola='OCU-1'),"
            "'preventiva', date('now','-1 year'), date('now','+45 days'))")
    entra(client, 'admin@a.it')

    entro_30 = client.get('/stampe/scadenze/manutenzioni?divisione_id=tutte&periodo=30g')
    assert 'OCU-1' not in testo_di(entro_30.data)

    entro_90 = client.get('/stampe/scadenze/manutenzioni?divisione_id=tutte&periodo=90g')
    assert 'OCU-1' in testo_di(entro_90.data)


def test_intervallo_di_date_libere_filtra_per_la_finestra_indicata(client, app, dati):
    """Stesso principio applicato a periodo=date: la finestra scelta deve
    davvero delimitare cosa compare nel prospetto, non bastare che sia una
    coppia di date coerente fra loro."""
    from models import execute
    with app.app_context():
        execute(
            "INSERT INTO manutenzioni (apparecchio_id,tipo,data_intervento,prossima_scadenza) "
            "VALUES ((SELECT id FROM apparecchi WHERE matricola='OCU-1'),"
            "'preventiva', date('now','-1 year'), date('now','+20 days'))")
    entra(client, 'admin@a.it')
    oggi = date.today()

    fuori_finestra = client.get(
        '/stampe/scadenze/manutenzioni?divisione_id=tutte&periodo=date'
        f'&da={oggi.isoformat()}&a={(oggi + timedelta(days=10)).isoformat()}')
    assert 'OCU-1' not in testo_di(fuori_finestra.data)

    dentro_finestra = client.get(
        '/stampe/scadenze/manutenzioni?divisione_id=tutte&periodo=date'
        f'&da={oggi.isoformat()}&a={(oggi + timedelta(days=30)).isoformat()}')
    assert 'OCU-1' in testo_di(dentro_finestra.data)


def test_intervallo_di_date_libere_rispetta_il_confine_inferiore(client, app, dati):
    """Una scadenza fra oggi e la data "Da" richiesta deve restare fuori dal
    prospetto: non e' scaduta (quindi non compare fra le scadute, che restano
    tutto cio' che precede oggi) e precede l'inizio della finestra richiesta
    (quindi non deve comparire nemmeno fra le prossime). Se la query ignora il
    confine inferiore e continua a usare "prossima_scadenza >= date('now')",
    questa scadenza compare comunque: e' proprio quello che il rilievo
    descrive come 'il campo Da viene validato e poi ignorato'."""
    from models import execute
    with app.app_context():
        execute(
            "INSERT INTO manutenzioni (apparecchio_id,tipo,data_intervento,prossima_scadenza) "
            "VALUES ((SELECT id FROM apparecchi WHERE matricola='CAR-1'),"
            "'preventiva', date('now','-1 year'), date('now','+5 days'))")
    entra(client, 'admin@a.it')
    oggi = date.today()

    risposta = client.get(
        '/stampe/scadenze/manutenzioni?divisione_id=tutte&periodo=date'
        f'&da={(oggi + timedelta(days=15)).isoformat()}'
        f'&a={(oggi + timedelta(days=40)).isoformat()}')
    assert risposta.status_code == 200
    assert 'CAR-1' not in testo_di(risposta.data)


def test_scadute_e_in_scadenza_sono_complementari_intorno_a_oggi(client, app, dati):
    """Le due sezioni del prospetto devono usare la stessa definizione di
    "oggi". Se il confine delle scadute fosse calcolato da SQLite in UTC
    (date('now')) e quello delle scelte rapide da Python in ora locale
    (datetime.now().date()), una scadenza di ieri potrebbe non soddisfare ne'
    "< ieri" ne' ">= oggi" e sparire da entrambe le sezioni senza alcun
    segnale. Una scadenza di ieri deve comparire fra le scadute, una di oggi
    deve comparire fra le in scadenza: nessuna delle due deve mancare."""
    # Le date sono calcolate con lo stesso orologio di _intervallo() (Python,
    # ora locale), non con date('now','...') di SQLite: e' proprio la
    # discrepanza fra i due orologi il difetto che questo test deve
    # dimostrare risolto, quindi il test non puo' fare affidamento su
    # entrambi per generare i propri dati.
    from models import execute
    oggi = date.today()
    ieri = oggi - timedelta(days=1)
    with app.app_context():
        execute(
            "INSERT INTO manutenzioni (apparecchio_id,tipo,data_intervento,prossima_scadenza) "
            "VALUES ((SELECT id FROM apparecchi WHERE matricola='OCU-1'),"
            "'preventiva', date('now','-1 year'), ?)", (ieri.isoformat(),))
        execute(
            "INSERT INTO manutenzioni (apparecchio_id,tipo,data_intervento,prossima_scadenza) "
            "VALUES ((SELECT id FROM apparecchi WHERE matricola='CAR-1'),"
            "'preventiva', date('now','-1 year'), ?)", (oggi.isoformat(),))
    entra(client, 'admin@a.it')

    risposta = client.get('/stampe/scadenze/manutenzioni?divisione_id=tutte&periodo=30g')
    assert risposta.status_code == 200
    testo = testo_di(risposta.data)
    assert 'OCU-1' in testo
    assert 'CAR-1' in testo
    # Non basta che compaiano da qualche parte: la scadenza di ieri deve
    # essere finita proprio fra le scadute e quella di oggi proprio fra le
    # in scadenza, non entrambe nella stessa sezione per un altro bug.
    assert 'Totale: 1 scadute, 1 in scadenza' in testo


def test_un_intervallo_libero_che_parte_nel_passato_non_duplica_le_righe(client, app, dati):
    """"Cosa scade in tutto il 2026" e' una richiesta legittima, e il form non
    impedisce di indicare una data "Da" gia' passata. Ma le scadute sono per
    definizione tutto cio' che precede oggi, indipendentemente dal periodo:
    se la sezione "in scadenza" partisse dalla data indicata dall'utente,
    ogni scadenza compresa fra quella data e oggi soddisferebbe entrambe le
    query e verrebbe stampata due volte, con un totale che ne dichiara due
    dove il record e' uno solo. Un documento consegnato a terzi con conteggi
    gonfiati e' peggio di un documento che non esce."""
    from models import execute
    oggi = date.today()
    with app.app_context():
        execute(
            "INSERT INTO manutenzioni (apparecchio_id,tipo,data_intervento,prossima_scadenza) "
            "VALUES ((SELECT id FROM apparecchi WHERE matricola='OCU-1'),"
            "'preventiva', date('now','-1 year'), ?)",
            ((oggi - timedelta(days=1)).isoformat(),))
    entra(client, 'admin@a.it')

    risposta = client.get(
        '/stampe/scadenze/manutenzioni?divisione_id=tutte&periodo=date'
        f'&da={(oggi - timedelta(days=30)).isoformat()}'
        f'&a={(oggi + timedelta(days=30)).isoformat()}')
    assert risposta.status_code == 200
    testo = testo_di(risposta.data)
    # Compare una volta sola, fra le scadute: e' li' che appartiene.
    assert testo.count('OCU-1') == 1
    assert 'Totale: 1 scadute, 0 in scadenza' in testo


FOGLIO = ('application/vnd.openxmlformats-officedocument'
          '.spreadsheetml.sheet')


def test_inventario_in_excel(client, dati):
    entra(client, 'admin@a.it')
    risposta = client.get('/stampe/inventario?divisione_id=tutte&formato=excel')
    assert risposta.status_code == 200
    assert risposta.mimetype == FOGLIO
    assert risposta.data[:2] == b'PK'   # xlsx e' un archivio zip


def test_scadenze_in_excel(client, dati):
    entra(client, 'admin@a.it')
    risposta = client.get('/stampe/scadenze/manutenzioni'
                          '?divisione_id=tutte&periodo=30g&formato=excel')
    assert risposta.mimetype == FOGLIO
    assert risposta.data[:2] == b'PK'


def test_lo_scope_vale_anche_per_excel(client, dati):
    entra(client, 'utente@a.it')
    risposta = client.get(f"/stampe/inventario?divisione_id={dati['d2']}&formato=excel")
    assert risposta.data[:2] != b'PK'


def _valori_foglio(dati_xlsx):
    """Tutti i valori delle celle del foglio attivo, appiattiti in un insieme
    di stringhe: comodo per verificare presenza/assenza di una matricola
    senza dover conoscere l'esatta cella in cui finisce."""
    wb = load_workbook(io.BytesIO(dati_xlsx))
    ws = wb.active
    return {str(cella.value) for riga in ws.iter_rows() for cella in riga
            if cella.value is not None}


def test_excel_inventario_contiene_solo_le_righe_in_ambito(client, dati):
    """Stessa garanzia di test_utente_vede_solo_la_propria_divisione_nell_inventario_generale,
    ma sul foglio Excel: leggere davvero il contenuto, non fermarsi al
    formato del file (b'PK' dimostra solo che e' uno zip)."""
    entra(client, 'utente@a.it')
    risposta = client.get('/stampe/inventario?divisione_id=tutte&formato=excel')
    assert risposta.status_code == 200
    valori = _valori_foglio(risposta.data)
    assert 'OCU-1' in valori
    assert 'CAR-1' not in valori
    assert 'DIS-DIV-1' not in valori
    # Il foglio porta le stesse colonne del PDF, descrizione compresa.
    assert 'Descrizione' in valori
    assert 'Autorefrattometro' in valori


def test_excel_inventario_admin_vede_tutta_la_struttura(client, dati):
    """Simmetrico al precedente: per admin l'ambito 'tutte' e' l'intera
    struttura (comprese le divisioni disattivate), non oltre."""
    entra(client, 'admin@a.it')
    risposta = client.get('/stampe/inventario?divisione_id=tutte&formato=excel')
    valori = _valori_foglio(risposta.data)
    assert 'OCU-1' in valori
    assert 'CAR-1' in valori
    assert 'DIS-DIV-1' in valori
    assert 'ALT-1' not in valori


def test_la_colonna_logo_esiste(app):
    from models import query_all
    with app.app_context():
        colonne = {r['name'] for r in query_all("PRAGMA table_info(strutture)")}
    assert 'logo_path' in colonne


def test_un_logo_inesistente_non_impedisce_la_stampa(client, dati, app):
    """Un logo che punta a un file cancellato dal disco non deve mai
    trasformarsi in un 500: il motore lo ignora e il PDF esce comunque
    (il try/except di ReportPDF.header() e' testato altrove, qui verifichiamo
    che il percorso arrivi fino in fondo senza rompere la stampa)."""
    from models import execute
    with app.app_context():
        execute("UPDATE strutture SET logo_path = ? WHERE id = ?",
                ('strutture/1/loghi/sparito.png', dati['s1']))
    entra(client, 'admin@a.it')
    risposta = client.get('/stampe/inventario?divisione_id=tutte')
    assert risposta.data.startswith(b'%PDF')


def test_admin_carica_il_logo_della_propria_struttura(client, dati, app):
    """Caso positivo: l'upload sulla propria struttura salva il percorso
    relativo sotto strutture/<id>/loghi/, l'unico prefisso che la rotta
    /uploads/<path> sa isolare per struttura."""
    from models import query_one
    entra(client, 'admin@a.it')
    risposta = client.post(
        f"/strutture/{dati['s1']}/logo",
        data={'logo': (io.BytesIO(b'contenuto-finto-png'), 'logo.png')},
        content_type='multipart/form-data')
    assert risposta.status_code == 302
    with app.app_context():
        struttura_a = query_one("SELECT logo_path FROM strutture WHERE id = ?", (dati['s1'],))
    assert struttura_a['logo_path'] is not None
    assert struttura_a['logo_path'].startswith(f"strutture/{dati['s1']}/loghi/")


def test_admin_non_puo_caricare_il_logo_di_un_altra_struttura(client, dati, app):
    """Un admin non deve poter modificare il logo di una struttura che non e'
    la propria, nemmeno passando l'id nell'URL. Non basta controllare che la
    risposta non sia un successo: bisogna dimostrare che il logo della
    struttura B e' rimasto invariato nel database, altrimenti un redirect
    'per caso' non proverebbe nulla."""
    from models import query_one
    entra(client, 'admin@a.it')  # admin della struttura s1
    risposta = client.post(
        f"/strutture/{dati['s2']}/logo",
        data={'logo': (io.BytesIO(b'contenuto-finto-png'), 'logo.png')},
        content_type='multipart/form-data')
    assert risposta.status_code == 302
    with app.app_context():
        struttura_b = query_one("SELECT logo_path FROM strutture WHERE id = ?", (dati['s2'],))
    assert struttura_b['logo_path'] is None


def test_excel_scadenze_separa_scadute_e_in_scadenza(client, app, dati):
    """Il foglio delle scadenze ha una colonna 'Stato': verifica che le
    scadute e le in-scadenza finiscano davvero nella riga giusta, non solo
    che compaiano da qualche parte nel file."""
    from models import execute
    oggi = date.today()
    ieri = oggi - timedelta(days=1)
    with app.app_context():
        execute(
            "INSERT INTO manutenzioni (apparecchio_id,tipo,data_intervento,prossima_scadenza) "
            "VALUES ((SELECT id FROM apparecchi WHERE matricola='OCU-1'),"
            "'preventiva', date('now','-1 year'), ?)", (ieri.isoformat(),))
        execute(
            "INSERT INTO manutenzioni (apparecchio_id,tipo,data_intervento,prossima_scadenza) "
            "VALUES ((SELECT id FROM apparecchi WHERE matricola='CAR-1'),"
            "'preventiva', date('now','-1 year'), ?)", (oggi.isoformat(),))
    entra(client, 'admin@a.it')

    risposta = client.get('/stampe/scadenze/manutenzioni'
                          '?divisione_id=tutte&periodo=30g&formato=excel')
    assert risposta.status_code == 200
    wb = load_workbook(io.BytesIO(risposta.data))
    ws = wb.active
    righe = list(ws.iter_rows(min_row=4, values_only=True))
    mappa_stato = {riga[3]: riga[0] for riga in righe}  # Matricola -> Stato
    assert mappa_stato['OCU-1'] == 'Scaduta'
    assert mappa_stato['CAR-1'] == 'In scadenza'
