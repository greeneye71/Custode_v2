"""Nessuna rotta deve servire dati a chi non ha uno scope.

Il difetto che questi test inchiodano non e' un errore di calcolo: e' un ramo
che restituisce "nessun filtro" invece di "nessun dato". Non si vede leggendo
una pagina che funziona, si vede solo chiedendola da un account senza scope.
"""
import io

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader
from werkzeug.security import generate_password_hash


def _testo_html(dati):
    """Le rotte HTML mostrano il dato cosi' com'e' nel markup: basta decodificare."""
    return dati.decode('utf-8', errors='replace')


def _testo_excel(dati):
    """Un xlsx e' uno zip: la stringa cercata non compare mai nei byte grezzi.
    Bisogna aprire il foglio e leggere i valori delle celle davvero."""
    wb = load_workbook(io.BytesIO(dati))
    ws = wb.active
    valori = {str(c.value) for riga in ws.iter_rows() for c in riga if c.value is not None}
    return '\n'.join(valori)


def _testo_pdf(dati):
    """Il testo di un PDF vive dentro stream di contenuto: va estratto, non
    cercato nei byte grezzi (vedi tests/test_report_service.py:testo_di)."""
    lettore = PdfReader(io.BytesIO(dati))
    return '\n'.join(pagina.extract_text() for pagina in lettore.pages)


@pytest.fixture
def due_strutture(app):
    """Struttura A con un admin che restera' orfano, struttura B con i dati
    che nessuno di A deve poter vedere."""
    from models import execute
    with app.app_context():
        a = execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Clinica A','A',1)").lastrowid
        b = execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Clinica B','B',1)").lastrowid
        da = execute("INSERT INTO divisioni (nome,codice,struttura_id) VALUES ('Oculistica','OCU',?)", (a,)).lastrowid
        db_ = execute("INSERT INTO divisioni (nome,codice,struttura_id) VALUES ('Cardiologia','CAR',?)", (b,)).lastrowid
        hash_pw = generate_password_hash('Passw0rd!')
        execute("INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,struttura_id,primo_accesso) "
                "VALUES ('admin@a.it',?,'A','A','admin',?,0)", (hash_pw, a))
        senza = execute(
            "INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,struttura_id,primo_accesso) "
            "VALUES ('nessuno@a.it',?,'N','N','utente',?,0)", (hash_pw, a)).lastrowid
        app_a = execute("INSERT INTO apparecchi (divisione_id,struttura_id,matricola,marca,modello,stato) "
                        "VALUES (?,?,'OCU-1','REXXAM','OZY','funzionante')", (da, a)).lastrowid
        app_b = execute("INSERT INTO apparecchi (divisione_id,struttura_id,matricola,marca,modello,stato) "
                        "VALUES (?,?,'SEGRETO-B','SIEMENS','Y1','funzionante')", (db_, b)).lastrowid
        for ap in (app_a, app_b):
            execute("INSERT INTO manutenzioni (apparecchio_id,tipo,data_intervento,prossima_scadenza) "
                    "VALUES (?,'preventiva',date('now','-1 year'),date('now','+30 days'))", (ap,))
            execute("INSERT INTO verifiche (apparecchio_id,data_verifica,prossima_scadenza,esito) "
                    "VALUES (?,date('now','-1 year'),date('now','+60 days'),'positivo')", (ap,))
    return {'a': a, 'b': b, 'senza': senza, 'app_b': app_b}


def entra(client, email):
    client.post('/login', data={'email': email, 'password': 'Passw0rd!'})


def orfana(app, struttura_id):
    """Riproduce lo stato che si crea disattivando o eliminando la struttura."""
    from models import execute
    with app.app_context():
        execute("UPDATE utenti SET struttura_id = NULL WHERE struttura_id = ?", (struttura_id,))


ROTTE = [
    # (url, estrattore) — ogni rotta va interrogata nel formato in cui parla
    # davvero: identita' per l'HTML, foglio letto per l'xlsx, testo estratto
    # per il PDF. La versione precedente cercava la stringa nei byte grezzi
    # della risposta per tutte le rotte: per l'xlsx (uno zip) e il PDF (testo
    # dentro stream di contenuto) l'asserzione passava qualunque cosa
    # facesse il codice, difetto compreso — un test che non prova nulla.
    ('/apparecchi', _testo_html),
    ('/manutenzioni', _testo_html),
    # Nota: la rotta e' registrata come '/scadenzario', non
    # '/manutenzioni/scadenzario' (il blueprint non ha url_prefix e questo
    # unico @manutenzioni_bp.route non ripete '/manutenzioni' come fanno gli
    # altri) — verificato con app.url_map.iter_rules(). E' un'inconsistenza
    # di naming preesistente e indipendente da questo fix; url_for('manutenzioni.
    # scadenzario') la risolve comunque correttamente, quindi l'app non ne
    # risente, ma un test che chiama l'URL sbagliato riceve sempre 404 e
    # "SEGRETO-B non c'e'" e' vero per il motivo sbagliato: usiamo l'URL reale.
    ('/scadenzario', _testo_html),
    ('/verifiche', _testo_html),
    ('/export/apparecchi/excel', _testo_excel),
    ('/export/apparecchi/pdf', _testo_pdf),
    # Le due export dello scadenzario non passavano da filtro_divisione(): al
    # posto suo avevano tre rami scritti a mano, e quello per il ruolo 'admin'
    # interrogava prossime_scadenze senza filtro. Un admin della struttura A
    # scaricava lo scadenzario di tutte le strutture. Le altre export non ne
    # soffrivano perche' chiamavano gia' il filtro condiviso.
    ('/export/scadenzario/excel', _testo_excel),
    ('/export/scadenzario/pdf', _testo_pdf),
]


@pytest.mark.parametrize('rotta,estrattore', ROTTE)
def test_admin_senza_struttura_non_ottiene_dati_altrui(client, app, due_strutture, rotta, estrattore):
    """Il caso concreto: la struttura dell'admin sparisce e lui resta senza
    scope. Non deve diventare un lasciapassare su tutte le altre."""
    entra(client, 'admin@a.it')
    orfana(app, due_strutture['a'])
    risposta = client.get(rotta, follow_redirects=True)
    assert 'SEGRETO-B' not in estrattore(risposta.data)


@pytest.mark.parametrize('rotta,estrattore', ROTTE)
def test_utente_senza_divisioni_non_ottiene_dati(client, app, due_strutture, rotta, estrattore):
    """Controprova sul ramo che gia' funziona: se questo fallisce, la
    correzione ha rotto il caso sano invece di sistemare quello guasto."""
    entra(client, 'nessuno@a.it')
    risposta = client.get(rotta, follow_redirects=True)
    testo = estrattore(risposta.data)
    assert 'SEGRETO-B' not in testo
    assert 'OCU-1' not in testo


def test_admin_con_struttura_vede_i_propri(client, due_strutture):
    """Il filtro deve restare permissivo dove e' giusto che lo sia: un test
    che verifica solo le negazioni passerebbe anche con 'AND 1=0' ovunque."""
    entra(client, 'admin@a.it')
    risposta = client.get('/apparecchi')
    assert b'OCU-1' in risposta.data
    assert b'SEGRETO-B' not in risposta.data


def test_admin_con_struttura_esporta_il_proprio_scadenzario(client, due_strutture):
    """Controprova sull'export appena messo sotto filtro: senza di questa,
    un 'AND 1=0' incondizionato passerebbe entrambi i test negativi."""
    entra(client, 'admin@a.it')
    testo = _testo_excel(client.get('/export/scadenzario/excel').data)
    assert 'OCU-1' in testo
    assert 'SEGRETO-B' not in testo


def test_apparecchio_accessibile_rifiuta_senza_struttura(app, due_strutture):
    """models.apparecchio_accessibile ha lo stesso difetto del filtro:
    'struttura_id = ? OR ? IS NULL' accetta qualunque apparecchio quando la
    struttura attiva e' None. E' il controllo che protegge i download degli
    allegati, quindi non basta correggere le liste."""
    from flask import g
    from models import apparecchio_accessibile, query_one
    with app.test_request_context():
        g.user = query_one("SELECT * FROM utenti WHERE email='admin@a.it'")
        g.struttura_id = None
        g.divisioni = []
        assert apparecchio_accessibile(due_strutture['app_b']) is None


def test_apparecchio_accessibile_lascia_passare_il_superadmin(app, due_strutture):
    """Un superadmin che non impersona ha struttura_id None per progetto: e'
    il suo stato normale, non un difetto. La correzione deve distinguerlo
    dagli altri ruoli invece di negare a tutti."""
    from flask import g
    from models import apparecchio_accessibile, execute, query_one
    with app.app_context():
        execute("INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,primo_accesso) "
                "VALUES ('super@x.it','x','S','S','superadmin',0)")
    with app.test_request_context():
        g.user = query_one("SELECT * FROM utenti WHERE email='super@x.it'")
        g.struttura_id = None
        g.divisioni = []
        assert apparecchio_accessibile(due_strutture['app_b']) is not None


def test_gli_utenti_orfani_vengono_disattivati_se_ambigui_su_piu_strutture(app):
    """Con due o piu' strutture attive l'ambiguita' e' reale: nessuno puo'
    indovinare a quale appartenesse un admin senza struttura, e disattivarlo
    resta la scelta prudente (dal Task 1 non vede piu' dati, ma continuava
    ad autenticarsi). I superadmin hanno struttura_id NULL per progetto e
    non vanno toccati; i tecnici nemmeno, perche' sono legati alle strutture
    da tecnici_strutture."""
    from models import execute, query_one, apply_schema_updates
    with app.app_context():
        execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Prima','P1',1)")
        execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Seconda','P2',1)")
        execute("INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,attivo) "
                "VALUES ('orfano@a.it','x','O','O','admin',1)")
        execute("INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,attivo) "
                "VALUES ('super@x.it','x','S','S','superadmin',1)")
        execute("INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,attivo) "
                "VALUES ('tec@x.it','x','T','T','tecnico',1)")

        apply_schema_updates()

        assert query_one("SELECT attivo FROM utenti WHERE email='orfano@a.it'")['attivo'] == 0
        assert query_one("SELECT attivo FROM utenti WHERE email='super@x.it'")['attivo'] == 1
        assert query_one("SELECT attivo FROM utenti WHERE email='tec@x.it'")['attivo'] == 1


def test_utente_orfano_resta_attivo_se_non_esiste_alcuna_struttura(app):
    """Se non esiste ancora nessuna struttura (es. installazione mai passata
    da migrate_v2_0.py) non c'e' un posto dove inquadrare l'utente:
    disattivarlo qui lo bloccherebbe fuori senza che nessuno abbia un posto
    dove riassegnarlo. Va lasciato attivo, in attesa che una struttura venga
    creata."""
    from models import execute, query_one, apply_schema_updates
    with app.app_context():
        execute("INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,attivo) "
                "VALUES ('orfano@nostruttura.it','x','O','O','admin',1)")

        apply_schema_updates()

        riga = query_one(
            "SELECT attivo, struttura_id FROM utenti WHERE email='orfano@nostruttura.it'"
        )
        assert riga['attivo'] == 1
        assert riga['struttura_id'] is None


def test_utente_orfano_viene_adottato_dall_unica_struttura_attiva(app):
    """Lo scenario dell'aggiornamento reale: un'installazione monostruttura
    ferma su uno schema v1.x (utenti senza struttura_id, CHECK a due soli
    ruoli) esegue apply_schema_updates(). La migrazione v2.2 aggiunge
    struttura_id lasciandolo NULL per l'admin esistente (in uno schema che
    non conosceva ancora superadmin/tecnico, l'unico utente possibile è
    admin o utente). Con un'unica struttura attiva non c'e' ambiguita' su
    dove inquadrarlo: deve restare attivo e ritrovarsi assegnato a quella
    struttura, non disattivato. E' esattamente il caso che una
    disattivazione incondizionata bloccherebbe fuori senza che nessuno
    abbia sbagliato niente."""
    from models import get_db, execute, query_one, apply_schema_updates
    with app.app_context():
        db = get_db()
        db.execute("PRAGMA foreign_keys = OFF")
        db.execute("DROP TABLE utenti")
        db.execute("""
            CREATE TABLE utenti (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              email TEXT UNIQUE NOT NULL,
              password_hash TEXT NOT NULL,
              nome TEXT NOT NULL,
              cognome TEXT NOT NULL,
              ruolo TEXT NOT NULL CHECK(ruolo IN ('admin', 'utente')),
              divisione_default_id INTEGER,
              attivo INTEGER DEFAULT 1,
              primo_accesso INTEGER DEFAULT 1,
              ultimo_accesso DATETIME,
              created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
              updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
              FOREIGN KEY (divisione_default_id) REFERENCES divisioni(id)
            )
        """)
        db.commit()

        struttura_id = execute(
            "INSERT INTO strutture (nome,codice,attiva) VALUES ('Unica','UNI',1)"
        ).lastrowid
        execute(
            "INSERT INTO utenti (email,password_hash,nome,cognome,ruolo) "
            "VALUES ('admin@unica.it','x','A','A','admin')"
        )

        apply_schema_updates()

        riga = query_one(
            "SELECT attivo, struttura_id FROM utenti WHERE email='admin@unica.it'"
        )
        assert riga['attivo'] == 1
        assert riga['struttura_id'] == struttura_id


def test_una_delete_diretta_non_puo_piu_orfanare_un_admin(app):
    """Il Task 1 impedisce all'orfano di vedere dati e il Task 2 lo disattiva,
    ma la condizione continuava a prodursi a ogni DELETE. Con RESTRICT la
    cancellazione viene rifiutata finche' gli utenti sono al loro posto: e'
    il Task 4 a cancellarli esplicitamente, prima della struttura."""
    import sqlite3
    import pytest
    from models import execute, get_db
    with app.app_context():
        s = execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Da Cancellare','DC',1)").lastrowid
        execute("INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,struttura_id) "
                "VALUES ('admin@dc.it','x','A','A','admin',?)", (s,))
        db = get_db()
        db.execute("PRAGMA foreign_keys = ON")
        with pytest.raises(sqlite3.IntegrityError):
            db.execute("DELETE FROM strutture WHERE id = ?", (s,))


def test_la_migrazione_non_rompe_le_fk_delle_tabelle_figlie(app):
    """Il modo in cui questa migrazione fallisce non e' un'eccezione: e' un
    database che sembra a posto e in cui sessioni punta a utenti_old_26. Si
    scopre al primo login, con un 500 'no such table'."""
    from models import get_db
    with app.app_context():
        db = get_db()
        for tabella in ('sessioni', 'utenti_divisioni', 'tecnici_strutture'):
            sql = db.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name=?",
                (tabella,)
            ).fetchone()[0]
            assert 'utenti_old' not in sql, f"{tabella} punta a una tabella rinominata"
        db.execute("PRAGMA foreign_keys = ON")
        assert db.execute("PRAGMA foreign_key_check").fetchall() == []


def test_migrazione_v26_scarta_colonne_assenti_nella_nuova_tabella(app):
    """cols_vecchie viene da PRAGMA table_info sulla utenti precedente: puo'
    contenere colonne che la nuova tabella non conosce (installazione
    personalizzata, o colonna rimossa fra le versioni). L'INSERT deve
    scartarle, non fallire: un dato che sparisce in silenzio e' accettabile
    per una colonna extra, un'app che non si avvia piu' no."""
    from models import get_db, execute, query_one, query_all, apply_schema_updates
    with app.app_context():
        db = get_db()
        db.execute("PRAGMA foreign_keys = OFF")
        db.execute("DROP TABLE utenti")
        db.execute("""
            CREATE TABLE utenti (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              email TEXT UNIQUE NOT NULL,
              password_hash TEXT NOT NULL,
              nome TEXT NOT NULL,
              cognome TEXT NOT NULL,
              ruolo TEXT NOT NULL CHECK(ruolo IN ('superadmin', 'admin', 'utente', 'tecnico')),
              divisione_default_id INTEGER,
              attivo INTEGER DEFAULT 1,
              primo_accesso INTEGER DEFAULT 1,
              ultimo_accesso DATETIME,
              created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
              updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
              struttura_id INTEGER,
              nota_personalizzata TEXT,
              FOREIGN KEY (struttura_id) REFERENCES strutture(id) ON DELETE SET NULL,
              FOREIGN KEY (divisione_default_id) REFERENCES divisioni(id) ON DELETE SET NULL
            )
        """)
        db.commit()

        s = execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Custom','CU',1)").lastrowid
        admin_id = execute(
            "INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,struttura_id,nota_personalizzata) "
            "VALUES ('admin@custom.it','x','A','A','admin',?,'campo che la 2.6 non conosce')",
            (s,)
        ).lastrowid
        utente_id = execute(
            "INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,struttura_id) "
            "VALUES ('utente@custom.it','x','U','U','utente',?)",
            (s,)
        ).lastrowid
        execute(
            "INSERT INTO sessioni (utente_id,token,expires_at) VALUES (?, 'tok-custom', datetime('now','+1 day'))",
            (admin_id,)
        )

        apply_schema_updates()

        assert query_one("SELECT COUNT(*) AS n FROM utenti")['n'] == 2
        assert query_all("SELECT name FROM sqlite_master WHERE name LIKE '%utenti_old%'") == []
        db.execute("PRAGMA foreign_keys = ON")
        assert db.execute("PRAGMA foreign_key_check").fetchall() == []
        sql_utenti = query_one("SELECT sql FROM sqlite_master WHERE name='utenti'")['sql']
        assert 'ON DELETE RESTRICT' in sql_utenti
        assert 'nota_personalizzata' not in sql_utenti
        # Scartare nota_personalizzata non basta: le colonne che restano devono
        # anche finire al posto giusto. Senza l'elenco delle colonne di
        # destinazione nell'INSERT i valori slittano, e struttura_id e' la prima
        # a raccogliere quello sbagliato — con l'effetto di spostare l'utente in
        # un'altra struttura, o fuori da tutte.
        assert query_one("SELECT struttura_id FROM utenti WHERE email='admin@custom.it'")['struttura_id'] == s
        assert query_one("SELECT struttura_id FROM utenti WHERE email='utente@custom.it'")['struttura_id'] == s
        riga = query_one(
            "SELECT s.token, u.email FROM sessioni s JOIN utenti u ON u.id = s.utente_id "
            "WHERE s.token = 'tok-custom'"
        )
        assert riga is not None and riga['email'] == 'admin@custom.it'
        _ = utente_id  # inserito per il conteggio, non serve altro riferimento


# ---------------------------------------------------------------------------
# Critico 1 (revisione finale 2.6): le rotte per-id di apparecchi.py,
# manutenzioni.py e verifiche.py avevano una copia inline del filtro di
# struttura — "AND (struttura_id = ? OR ? IS NULL)" — invece di passare da
# models.apparecchio_accessibile(). Con g.struttura_id None quella copia
# accetta qualunque riga di qualunque struttura. Si raggiunge senza alcuno
# stato anomalo: un tecnico assegnato a piu' strutture, subito dopo il login
# (prima che scelga quale impersonare), oppure un admin la cui struttura e'
# stata disattivata (il login controlla solo utenti.attivo, non lo stato
# della struttura). I due scenari sotto riproducono esattamente questi casi.
# ---------------------------------------------------------------------------

def disattiva_struttura(app, struttura_id):
    """Scenario 2: la struttura viene disattivata (es. cessata l'attivita'),
    ma il login (auth.py, query 'WHERE email=? AND attivo=1' sugli utenti)
    non verifica lo stato della struttura dell'utente. L'admin resta
    autenticato; in _load_user_from_session il ramo che risolve la struttura
    non trova nulla ('WHERE id=? AND attiva=1') e g.struttura_id diventa
    None — esattamente come per l'admin orfano di 'due_strutture', ma per
    una via diversa e altrettanto raggiungibile."""
    from models import execute
    with app.app_context():
        execute("UPDATE strutture SET attiva = 0 WHERE id = ?", (struttura_id,))


@pytest.fixture
def tecnico_appena_entrato(app):
    """Scenario 1: un tecnico assegnato a due strutture (A e B). Subito dopo
    il login, prima che scelga quale impersonare, la sessione non ha
    struttura_impersonata_id: g.struttura_id e' None. I dati bersaglio vivono
    in una terza struttura (C), a cui il tecnico non e' affatto assegnato —
    cosi' il test dimostra che l'assenza di scope si comporta da lasciapassare
    universale, non solo verso A o B."""
    from models import execute
    with app.app_context():
        a = execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Clinica A','TA',1)").lastrowid
        b = execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Clinica B','TB',1)").lastrowid
        c = execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Clinica C','TC',1)").lastrowid
        dc = execute("INSERT INTO divisioni (nome,codice,struttura_id) VALUES ('Div C','DVC',?)", (c,)).lastrowid
        hash_pw = generate_password_hash('Passw0rd!')
        tecnico_id = execute(
            "INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,primo_accesso) "
            "VALUES ('tecnico@t.it',?,'T','T','tecnico',0)", (hash_pw,)
        ).lastrowid
        execute("INSERT INTO tecnici_strutture (tecnico_id,struttura_id) VALUES (?,?)", (tecnico_id, a))
        execute("INSERT INTO tecnici_strutture (tecnico_id,struttura_id) VALUES (?,?)", (tecnico_id, b))
        app_c = execute(
            "INSERT INTO apparecchi (divisione_id,struttura_id,matricola,marca,modello,stato) "
            "VALUES (?,?,'SEGRETO-C','SIEMENS','C1','funzionante')", (dc, c)
        ).lastrowid
        manutenzione_c = execute(
            "INSERT INTO manutenzioni (apparecchio_id,tipo,data_intervento,prossima_scadenza,descrizione) "
            "VALUES (?,'preventiva',date('now','-1 year'),date('now','+30 days'),'SEGRETO-C-MANUTENZIONE')",
            (app_c,)
        ).lastrowid
        verifica_c = execute(
            "INSERT INTO verifiche (apparecchio_id,data_verifica,prossima_scadenza,esito,note) "
            "VALUES (?,date('now','-1 year'),date('now','+60 days'),'positivo','SEGRETO-C-VERIFICA')",
            (app_c,)
        ).lastrowid
    return {
        'a': a, 'b': b, 'c': c, 'app_c': app_c,
        'manutenzione_c': manutenzione_c, 'verifica_c': verifica_c,
    }


def test_tecnico_appena_entrato_va_alla_scelta_struttura(client, tecnico_appena_entrato):
    """Precondizione dello scenario 1, dimostrata dalla revisione: il login
    di un tecnico con piu' strutture assegnate rediretta alla scelta, non
    imposta alcuna struttura di default."""
    risposta = client.post('/login', data={'email': 'tecnico@t.it', 'password': 'Passw0rd!'})
    assert risposta.status_code == 302
    assert '/tecnico/seleziona-struttura' in risposta.headers['Location']


def test_tecnico_appena_entrato_non_legge_apparecchio_di_struttura_estranea(client, tecnico_appena_entrato):
    """Lettura (apparecchi.py, dettaglio): senza scope, la rotta deve
    comportarsi come se l'apparecchio non esistesse, non mostrarlo."""
    entra(client, 'tecnico@t.it')
    risposta = client.get(f"/apparecchi/{tecnico_appena_entrato['app_c']}")
    assert risposta.status_code == 302


def test_tecnico_appena_entrato_non_dismette_apparecchio_di_struttura_estranea(client, app, tecnico_appena_entrato):
    """Cancellazione (apparecchi.py, dismetti): senza scope, il tecnico non
    deve poter dismettere un apparecchio di una struttura a cui non e'
    nemmeno assegnato."""
    from models import query_one
    entra(client, 'tecnico@t.it')
    client.post(f"/apparecchi/{tecnico_appena_entrato['app_c']}/dismetti")
    with app.app_context():
        riga = query_one("SELECT stato FROM apparecchi WHERE id=?", (tecnico_appena_entrato['app_c'],))
    assert riga['stato'] == 'funzionante'


def test_tecnico_appena_entrato_non_legge_manutenzione_di_struttura_estranea(client, tecnico_appena_entrato):
    """Lettura (manutenzioni.py, modifica GET): la riga appartiene a un
    apparecchio di una struttura estranea, il cancello e' l'apparecchio."""
    entra(client, 'tecnico@t.it')
    risposta = client.get(f"/manutenzioni/{tecnico_appena_entrato['manutenzione_c']}/modifica")
    assert risposta.status_code == 302


def test_tecnico_appena_entrato_non_legge_verifica_di_struttura_estranea(client, tecnico_appena_entrato):
    """Lettura (verifiche.py, modifica GET): stessa forma della precedente,
    file gemello."""
    entra(client, 'tecnico@t.it')
    risposta = client.get(f"/verifiche/{tecnico_appena_entrato['verifica_c']}/modifica")
    assert risposta.status_code == 302


def test_admin_struttura_disattivata_non_modifica_apparecchio_altrui(client, app, due_strutture):
    """Scrittura (apparecchi.py, modifica POST): e' l'esploit dimostrato dalla
    revisione — 'SCRITTO DA A' diventava 'RISCRITTO' sulla riga della
    Clinica B. Qui verifichiamo che la marca non cambi."""
    from models import query_one
    entra(client, 'admin@a.it')
    disattiva_struttura(app, due_strutture['a'])
    with app.app_context():
        prima = query_one("SELECT divisione_id FROM apparecchi WHERE id=?", (due_strutture['app_b'],))
    client.post(f"/apparecchi/{due_strutture['app_b']}/modifica", data={
        'matricola': 'SEGRETO-B',
        'marca': 'RISCRITTO',
        'modello': 'Y1',
        'divisione_id': str(prima['divisione_id']),
    })
    with app.app_context():
        dopo = query_one("SELECT marca FROM apparecchi WHERE id=?", (due_strutture['app_b'],))
    assert dopo['marca'] == 'SIEMENS'


def test_admin_struttura_disattivata_non_scrive_manutenzione_altrui(client, app, due_strutture):
    """Scrittura (manutenzioni.py, modifica POST): senza scope la rotta deve
    rifiutare subito (redirect), non arrivare fino alla validazione del
    form — e in nessun caso deve modificare la riga."""
    from models import query_one
    entra(client, 'admin@a.it')
    disattiva_struttura(app, due_strutture['a'])
    with app.app_context():
        manutenzione_b = query_one(
            "SELECT id, tipo FROM manutenzioni WHERE apparecchio_id=?", (due_strutture['app_b'],)
        )
    risposta = client.post(f"/manutenzioni/{manutenzione_b['id']}/modifica", data={
        'apparecchio_id': str(due_strutture['app_b']),
        'tipo': 'correttiva',
        'data_intervento': '2026-01-01',
    })
    assert risposta.status_code == 302
    with app.app_context():
        dopo = query_one("SELECT tipo FROM manutenzioni WHERE id=?", (manutenzione_b['id'],))
    assert dopo['tipo'] == manutenzione_b['tipo']


def test_admin_struttura_disattivata_non_elimina_manutenzione_altrui(client, app, due_strutture):
    """Cancellazione (manutenzioni.py, elimina POST): e' l'esploit dimostrato
    dalla revisione — la manutenzione della Clinica B veniva cancellata da un
    admin della Clinica A senza scope. La DELETE non ha un proprio filtro di
    struttura: e' il cancello iniziale a dover fermarla."""
    from models import query_one
    entra(client, 'admin@a.it')
    disattiva_struttura(app, due_strutture['a'])
    with app.app_context():
        manutenzione_b = query_one(
            "SELECT id FROM manutenzioni WHERE apparecchio_id=?", (due_strutture['app_b'],)
        )
    client.post(f"/manutenzioni/{manutenzione_b['id']}/elimina")
    with app.app_context():
        esiste = query_one("SELECT id FROM manutenzioni WHERE id=?", (manutenzione_b['id'],))
    assert esiste is not None


def test_admin_struttura_disattivata_non_scrive_verifica_altrui(client, app, due_strutture):
    """Scrittura (verifiche.py, modifica POST): file gemello di manutenzioni,
    stessa forma di test."""
    from models import query_one
    entra(client, 'admin@a.it')
    disattiva_struttura(app, due_strutture['a'])
    with app.app_context():
        verifica_b = query_one(
            "SELECT id, esito FROM verifiche WHERE apparecchio_id=?", (due_strutture['app_b'],)
        )
    risposta = client.post(f"/verifiche/{verifica_b['id']}/modifica", data={
        'apparecchio_id': str(due_strutture['app_b']),
        'data_verifica': '2026-01-01',
        'esito': 'negativo',
    })
    assert risposta.status_code == 302
    with app.app_context():
        dopo = query_one("SELECT esito FROM verifiche WHERE id=?", (verifica_b['id'],))
    assert dopo['esito'] == verifica_b['esito']


def test_admin_struttura_disattivata_non_elimina_verifica_altrui(client, app, due_strutture):
    """Cancellazione (verifiche.py, elimina POST): e' l'esploit dimostrato
    dalla revisione — la verifica della Clinica B veniva cancellata da un
    admin della Clinica A senza scope."""
    from models import query_one
    entra(client, 'admin@a.it')
    disattiva_struttura(app, due_strutture['a'])
    with app.app_context():
        verifica_b = query_one(
            "SELECT id FROM verifiche WHERE apparecchio_id=?", (due_strutture['app_b'],)
        )
    client.post(f"/verifiche/{verifica_b['id']}/elimina")
    with app.app_context():
        esiste = query_one("SELECT id FROM verifiche WHERE id=?", (verifica_b['id'],))
    assert esiste is not None
